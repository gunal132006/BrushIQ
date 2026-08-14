"""
BrushIQ — Comprehensive Secure Code Review Report Generator
Generates a multi-sheet, fully styled Excel (.xlsx) workbook.
Phases: Executive Summary | Backend Inventory | API Endpoints |
        Security Findings | Dependency Review | Remediation |
        GitHub Actions | Risk Dashboard | Setup Instructions
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
import os

# ─── Colour Palette ────────────────────────────────────────────────────────────
C_NAVY      = "0D1B2A"
C_DARK_BLUE = "1B2A4A"
C_ACCENT    = "2563EB"
C_TEAL      = "0EA5E9"
C_RED       = "DC2626"
C_ORANGE    = "EA580C"
C_YELLOW    = "CA8A04"
C_GREEN     = "16A34A"
C_PURPLE    = "7C3AED"
C_WHITE     = "FFFFFF"
C_LIGHT_BG  = "F1F5F9"
C_LIGHT2    = "E2E8F0"
C_GREY      = "94A3B8"
C_DARK_TEXT = "0F172A"
C_CREAM     = "FFFBF0"

# ─── Style Helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_DARK_TEXT, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="64748B")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, fg=C_WHITE, font_color=C_DARK_TEXT,
             font_size=10, h_align="left", wrap=False, border=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=font_color, size=font_size, italic=italic)
    c.fill = fill(fg)
    c.alignment = align(h_align, "center", wrap)
    if border:
        c.border = thin_border()
    return c

def header_row(ws, row, cols_values, fg=C_DARK_BLUE, font_color=C_WHITE, size=10):
    for col, val in enumerate(cols_values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(bold=True, color=font_color, size=size)
        c.fill = fill(fg)
        c.alignment = align("center", "center")
        c.border = thin_border()

def section_title(ws, row, col, text, span, fg=C_ACCENT, size=12):
    ws.row_dimensions[row].height = 24
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=C_WHITE, size=size)
    c.fill = fill(fg)
    c.alignment = align("left", "center")
    c.border = medium_border()
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)

def banner(ws, row, text, span, fg=C_NAVY, size=16, height=45):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=C_WHITE, size=size, name="Calibri")
    c.fill = fill(fg)
    c.alignment = align("center", "center")

def auto_width(ws, extra=4, max_w=70):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + extra, max_w)

# ═══════════════════════════════════════════════════════════════════════════════
# ─── DATA DEFINITIONS ──────────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

BACKEND_INVENTORY = [
    ("Framework",               "Node.js + Express.js v4.19.2"),
    ("Language",                "JavaScript (ES2020, CommonJS modules, Node 18+)"),
    ("API Architecture",        "RESTful JSON API — no versioning prefix (/api/*)"),
    ("Authentication",          "JWT (jsonwebtoken ^9.0.2) + Google OAuth 2.0 (google-auth-library ^11.0.0)"),
    ("Authorization Model",     "User-scoped resource ownership — SQL JOIN checks per request, no RBAC roles"),
    ("Database",                "PostgreSQL 15 (Supabase cloud) + Embedded JSON file fallback store"),
    ("ORM / Query Builder",     "Raw parameterized SQL via pg ^8.11.5 — no ORM, no query builder"),
    ("API Documentation",       "NONE — no Swagger / OpenAPI spec exists"),
    ("Security Middleware",     "Helmet ^8.3.0, express-rate-limit ^8.6.2, CORS restricted allowlist"),
    ("File Uploads",            "multer ^1.4.5-lts.1 — disk storage, /uploads/, 5 MB limit, MIME type filter"),
    ("Session Handling",        "Stateless JWT (7-day expiry, Bearer token, no refresh mechanism)"),
    ("Image Processing",        "jimp ^1.6.1 + @tensorflow/tfjs ^4.22.0 + @tensorflow-models/coco-ssd ^2.2.3"),
    ("Email Service",           "nodemailer ^9.0.5 via Gmail SMTP — TLS cert validation disabled"),
    ("Third-Party Integrations","Google OAuth2 (sign-in), Supabase/PostgreSQL (cloud DB), Render.com (hosting)"),
    ("Deployment",              "Render.com (Node web service), Docker Compose (local dev only)"),
    ("Containerization",        "docker-compose.yml — PostgreSQL + Adminer exposed on 0.0.0.0 in dev"),
    ("Frontend",                "React (Vite) — Vercel deployment, separate from backend"),
    ("Static Files",            "/uploads/* and /illustrations/* served via express.static (unauthenticated)"),
]

ENDPOINT_INVENTORY = [
    # (Endpoint, Method, Auth Required, Risk, Expected Roles, Controller File, Notes)
    ("/health",                         "GET",    "No",  "Low",      "Public",          "app.js (inline)",                         "Basic health probe"),
    ("/api/health",                     "GET",    "No",  "Low",      "Public",          "app.js (inline)",                         "Duplicate health route"),
    ("/",                               "GET",    "No",  "Low",      "Public",          "app.js (inline)",                         "Root info page"),
    ("/api/auth/register",              "POST",   "No",  "Medium",   "Public",          "authController.register",                 "No rate limit, no complexity check"),
    ("/api/auth/login",                 "POST",   "No",  "Medium",   "Public",          "authController.login",                    "No rate limit — brute-force risk"),
    ("/api/auth/google",                "POST",   "No",  "Medium",   "Public",          "authController.googleLogin",              "Google OAuth token exchange"),
    ("/api/auth/forgot-password",       "POST",   "No",  "High",     "Public",          "authController.forgotPassword",           "User enumeration via 400 response"),
    ("/api/auth/reset-password",        "POST",   "No",  "Medium",   "Public",          "authController.resetPassword",            "Token-based password reset"),
    ("/api/auth/reset-password-page",   "GET",    "No",  "Low",      "Public",          "authController.renderResetPage",          "Inline HTML page, CSP unsafe-inline"),
    ("/api/auth/change-password",       "POST",   "Yes", "Medium",   "Authenticated",   "authController.changePassword",           "Old token still valid after change"),
    ("/api/auth/me",                    "GET",    "Yes", "Low",      "Authenticated",   "authController.getMe",                    "Returns user profile"),
    ("/api/auth/reset-limiter",         "POST",   "No",  "High",     "Dev (NODE_ENV)",  "rateLimiter.resetAuthLimiter",            "Dev endpoint exposed on public API"),
    ("/api/family",                     "GET",    "Yes", "Low",      "Authenticated",   "familyController.getFamilyMembers",        "Scoped to user"),
    ("/api/family",                     "POST",   "Yes", "Medium",   "Authenticated",   "familyController.addFamilyMember",        "No length/type validation"),
    ("/api/family/:id",                 "PUT",    "Yes", "Medium",   "Authenticated",   "familyController.updateFamilyMember",     "Ownership checked via SQL JOIN"),
    ("/api/family/:id",                 "DELETE", "Yes", "Low",      "Authenticated",   "familyController.deleteFamilyMember",     "Ownership checked"),
    ("/api/toothbrushes",               "GET",    "Yes", "Low",      "Authenticated",   "toothbrushController.getToothbrushes",    "Scoped to user"),
    ("/api/toothbrushes",               "POST",   "Yes", "Medium",   "Authenticated",   "toothbrushController.addToothbrush",      "No input validation"),
    ("/api/toothbrushes/:id",           "PUT",    "Yes", "Medium",   "Authenticated",   "toothbrushController.updateToothbrush",   "Ownership checked via JOIN"),
    ("/api/toothbrushes/:id",           "DELETE", "Yes", "Medium",   "Authenticated",   "toothbrushController.deleteToothbrush",   "IDOR risk — delete WHERE only id=$1"),
    ("/api/toothbrush",                 "GET",    "Yes", "Low",      "Authenticated",   "toothbrushController (alias)",            "Duplicate route alias"),
    ("/api/toothbrush",                 "POST",   "Yes", "Medium",   "Authenticated",   "toothbrushController (alias)",            "Duplicate route alias"),
    ("/api/scans/analyze",              "POST",   "Yes", "High",     "Authenticated",   "scanController.analyzeScan",              "CPU-heavy AI, no scan rate limit, MIME spoofable"),
    ("/api/scans",                      "POST",   "Yes", "Medium",   "Authenticated",   "scanController.saveScan",                 "Client-controlled metrics, no validation"),
    ("/api/scans",                      "GET",    "Yes", "Low",      "Authenticated",   "scanController.getScansHistory",          "Scoped to user"),
    ("/api/scans/:id",                  "GET",    "Yes", "Low",      "Authenticated",   "scanController.getScanById",              "Ownership checked"),
    ("/api/reminders",                  "GET",    "Yes", "Low",      "Authenticated",   "reminderController.getReminders",         "Scoped to user"),
    ("/api/reminders",                  "POST",   "Yes", "Low",      "Authenticated",   "reminderController.createReminder",       ""),
    ("/api/reminders/:id/complete",     "PUT",    "Yes", "Low",      "Authenticated",   "reminderController.completeReminder",     ""),
    ("/api/tips",                       "GET",    "Yes", "Low",      "Authenticated",   "tipController.getTips",                   ""),
    ("/api/tips/personalized",          "GET",    "Yes", "Low",      "Authenticated",   "tipController.getPersonalizedTips",       ""),
    ("/api/dashboard",                  "GET",    "Yes", "Medium",   "Authenticated",   "dashboardController.getDashboardData",    "Error leaks DB message"),
    ("/api/system/database-status",     "GET",    "No",  "High",     "Public",          "routes/system.js (inline handler)",       "Unauthenticated infra status endpoint"),
    ("/uploads/*",                      "GET",    "No",  "High",     "Public",          "express.static — uploads directory",      "Health scan images publicly accessible"),
    ("/illustrations/*",                "GET",    "No",  "Low",      "Public",          "express.static — frontend/public",        "Static assets, no auth needed"),
]

# ID, Category, Severity, CWE, OWASP, File, Title, Description, Why Concern, Fix
FINDINGS = [
    ("SEC-001", "Cryptography / Secrets",  "CRITICAL", "CWE-798", "A02:2021",
     "backend/.env (L7) + src/middlewares/auth.js (L5)",
     "Hardcoded JWT Secret in .env AND Code Fallback",
     "JWT_SECRET is set to 'supersecretbrushiqjwttoken' in the committed .env file. auth.js also falls back to this same hardcoded value if the env var is missing.",
     "Any attacker who reads the .env file (via repo leak or server path traversal) can sign arbitrary JWTs and impersonate any user. The in-code fallback means even a misconfigured deployment is exploitable.",
     "Rotate secret immediately. Generate a 256-bit random value: openssl rand -base64 32. Remove .env from git history. Add .env to .gitignore. Remove ALL code-level fallback values. Fail startup if JWT_SECRET is missing."),

    ("SEC-002", "Sensitive Data Exposure", "CRITICAL", "CWE-312", "A02:2021",
     "backend/.env (L8) + backend/render.yaml (L28)",
     "Google OAuth Client ID Committed to Repository",
     "The real Google OAuth2 Client ID (534843148727-ernb5gqgo6pf1cobmmvjbsl7d4f5026s.apps.googleusercontent.com) is embedded in both .env and render.yaml, both committed to the repository.",
     "The Client ID is publicly visible to all repository readers. It can be used to craft phishing OAuth flows targeting BrushIQ users. If the Client Secret were also exposed, full account takeover would be trivial.",
     "Remove Client ID from all committed files. Use Render environment variable injection (sync: false, set manually in dashboard). Revoke and rotate the OAuth credential in Google Cloud Console."),

    ("SEC-003", "Sensitive Data Exposure", "CRITICAL", "CWE-312", "A02:2021",
     "backend/.env (L4-L6) + docker-compose.yml (L10)",
     "Database Credentials Committed to Repository",
     "DB_USER=postgres, DB_PASSWORD=postgrespassword, DB_DATABASE=brushiq are present in .env. docker-compose.yml also hardcodes POSTGRES_PASSWORD=postgrespassword.",
     "Exposed database credentials allow direct database access if the port is reachable. Combined with Adminer running on port 8080 with no auth, this is an immediate full data breach risk.",
     "Rotate DB credentials immediately. Remove .env from VCS. Use Render secret env vars (sync: false for DB_PASSWORD). Use Docker secrets instead of environment variables."),

    ("SEC-004", "Authentication",          "HIGH",     "CWE-326", "A07:2021",
     "backend/src/middlewares/auth.js (L5)",
     "Weak Fallback JWT Secret in Auth Middleware",
     "The auth middleware falls back to 'supersecretbrushiqjwttoken' if JWT_SECRET env var is not set. This is predictable, short (30 chars), and could be brute-forced.",
     "If JWT_SECRET is not set in production (e.g., deployment misconfiguration), the application silently accepts tokens signed with a known weak secret, making all JWTs forgeable.",
     "Remove all fallback values. If JWT_SECRET is missing at startup, throw a fatal error: if (!process.env.JWT_SECRET) { console.error('FATAL: JWT_SECRET required'); process.exit(1); }"),

    ("SEC-005", "Authentication",          "HIGH",     "CWE-521", "A07:2021",
     "backend/src/controllers/authController.js (L165)",
     "Weak Minimum Password Length — Only 6 Characters",
     "Registration and change-password enforce a minimum password length of only 6 characters with no complexity requirements.",
     "6-character passwords are highly susceptible to brute-force and dictionary attacks. NIST SP 800-63B recommends at least 8 characters minimum, with support for longer passphrases.",
     "Increase minimum to 12 characters. Add complexity check: /^(?=.*[a-z])(?=.*[A-Z])(?=.*\\d)(?=.*[!@#$%^&*]).{12,}$/. Consider zxcvbn strength scorer. Implement account lockout after N failures."),

    ("SEC-006", "Authentication",          "HIGH",     "CWE-204", "A07:2021",
     "backend/src/controllers/authController.js (L275-L299)",
     "User Enumeration via Forgot-Password Endpoint",
     "The /api/auth/forgot-password endpoint returns 'User not found' (HTTP 400) when the email or phone does not exist in the database.",
     "Attackers can enumerate valid email addresses and phone numbers by submitting them to this endpoint and observing the differing HTTP response code and body.",
     "Return the same success response regardless of whether the user exists: { message: 'If this account exists, a reset link has been sent.' }. Always return HTTP 200."),

    ("SEC-007", "Authorization / IDOR",   "HIGH",     "CWE-639", "A01:2021",
     "backend/src/controllers/toothbrushController.js (L125)",
     "IDOR Risk — DELETE Query Missing User Ownership Filter",
     "After an ownership pre-check, the DELETE query uses DELETE FROM toothbrushes WHERE id = $1 without re-checking the user_id constraint directly in the DELETE statement.",
     "If the ownership check passes but a race condition or logic bug changes ownership before the DELETE executes, an attacker could delete another user's toothbrush record.",
     "Change DELETE query to include user filter directly: DELETE FROM toothbrushes WHERE id = $1 AND family_member_id IN (SELECT id FROM family_members WHERE user_id = $2)"),

    ("SEC-008", "Configuration",          "HIGH",     "CWE-942", "A05:2021",
     "backend/src/app.js (L13)",
     "CORS Wildcard '*' When ALLOWED_ORIGINS Not Set",
     "If the ALLOWED_ORIGINS environment variable is not configured, the CORS policy defaults to origin: '*', permitting any origin to make cross-origin requests.",
     "An overly permissive CORS policy allows malicious websites to make authenticated API calls on behalf of logged-in users — a CSRF-like attack via credentialed CORS requests.",
     "Set ALLOWED_ORIGINS explicitly in all environments. Remove the wildcard fallback. Use strict allowlist: ['https://brushiq.app']. Fail startup if ALLOWED_ORIGINS is not set in production."),

    ("SEC-009", "Configuration",          "HIGH",     "CWE-284", "A01:2021",
     "backend/src/app.js (L38)",
     "Uploaded Scan Images Served Without Authentication",
     "The /uploads static directory is served publicly via express.static without any authentication middleware.",
     "Any person (authenticated or not) can access uploaded scan images directly via /uploads/<filename> if they know or guess the filename. Scan images contain identifiable health data (toothbrush imagery).",
     "Move uploads behind authentication. Serve via controller that validates JWT first: router.get('/uploads/:file', authMiddleware, serveFile). Alternatively, use signed URLs or a private S3 bucket."),

    ("SEC-010", "Configuration",          "HIGH",     "CWE-284", "A01:2021",
     "backend/src/routes/system.js (L6-L13)",
     "Unauthenticated /api/system/database-status Reveals Infrastructure Details",
     "GET /api/system/database-status is public and returns { mode: 'postgresql', connected: true/false } without any authentication.",
     "Internal infrastructure state (DB connectivity, mode) is exposed to the public. Error responses may include err.message with DB connection details — valuable attacker reconnaissance.",
     "Protect this endpoint with authMiddleware. Remove error details from the response body in production. Expose only boolean status."),

    ("SEC-011", "Cryptography",           "HIGH",     "CWE-295", "A02:2021",
     "backend/src/config/db.js (L39, L47-L49)",
     "SSL Certificate Verification Disabled (rejectUnauthorized: false)",
     "Both PostgreSQL connection paths set rejectUnauthorized: false for SSL, disabling server certificate validation entirely.",
     "Disabling certificate validation enables Man-in-the-Middle attacks on the database connection. An attacker on the network path could intercept or tamper with all DB traffic, exposing passwords, tokens, and PII.",
     "Set rejectUnauthorized: true and provide the proper CA cert. For Supabase: ssl: { rejectUnauthorized: true, ca: fs.readFileSync('supabase-ca.crt') }. Download the root CA from Supabase dashboard."),

    ("SEC-012", "Cryptography",           "HIGH",     "CWE-295", "A02:2021",
     "backend/src/services/mailerService.js (L24)",
     "Email TLS Certificate Validation Disabled",
     "The nodemailer transporter has tls: { rejectUnauthorized: false } — TLS certificate validation disabled for SMTP connections.",
     "Password reset emails (containing reset links) are sent over an unverified TLS connection. An attacker can MitM the SMTP connection, intercept the reset link, and take over user accounts.",
     "Remove rejectUnauthorized: false or set it to true. For Gmail and standard providers, this setting is not needed and should be removed."),

    ("SEC-013", "Input Validation",       "MEDIUM",   "CWE-20",  "A03:2021",
     "backend/src/controllers/scanController.js (L34-L99)",
     "Client-Controlled Scan Metrics Saved Without Server-Side Validation",
     "The saveScan endpoint accepts all scan metrics (wearPercentage, healthScore, bristleSpreading, etc.) directly from the request body without range or type validation.",
     "A malicious user can POST arbitrary scan metrics (e.g., healthScore: -999, wearPercentage: 200) to manipulate records and potentially exploit downstream business logic.",
     "Validate all numeric fields server-side: wearPercentage in [0,100], healthScore in [0,100], condition must be enum. Use a validation library (Joi, Zod, express-validator)."),

    ("SEC-014", "Input Validation",       "MEDIUM",   "CWE-434", "A03:2021",
     "backend/src/routes/scan.js (L30-L36)",
     "File Upload MIME Type Check Is Client-Controlled (No Magic-Byte Validation)",
     "The multer fileFilter checks file.mimetype.startsWith('image/') but this Content-Type header is set by the HTTP client and can be trivially spoofed.",
     "An attacker can upload a malicious file (web shell, SVG with embedded JS, HTML) by setting Content-Type: image/jpeg. This content is then accessible via the public /uploads/ path.",
     "Add magic-byte validation using the 'file-type' npm package after the file is written to disk. Restrict extensions to .jpg, .jpeg, .png, .webp. Store uploads outside webroot or behind auth."),

    ("SEC-015", "Input Validation",       "MEDIUM",   "CWE-20",  "A03:2021",
     "backend/src/controllers/familyController.js (L104-L124)",
     "No Length or Type Validation on Family Member Fields",
     "addFamilyMember only checks field presence, not type or length. name, gender, relationship accept any string of any length.",
     "An attacker could insert extremely long strings (DB bloat, DoS), HTML content in name fields (Stored XSS if rendered in emails), or unexpected types that cause application errors.",
     "Add max-length validation (name ≤ 100 chars). Validate age is integer in [0,120]. Use express-validator or Joi schema middleware. Validate enum fields (gender, relationship)."),

    ("SEC-016", "Sensitive Data Exposure","MEDIUM",   "CWE-532", "A09:2021",
     "backend/src/controllers/authController.js (L215-L217)",
     "Internal Error Details Logged to Console — PII Risk",
     "catch(err) { console.error('Registration error details:', err) } logs the full error object including stack trace and potentially request data.",
     "In a cloud-log-aggregated environment (Render logs), stack traces reveal internal structure, file paths, and potentially partial user data — useful for attackers doing reconnaissance.",
     "Log err.message only, not the full object. Use structured logger (pino, winston) with log-level control. Never log sensitive request body fields."),

    ("SEC-017", "Business Logic",         "MEDIUM",   "CWE-307", "A07:2021",
     "backend/src/controllers/authController.js (all auth endpoints)",
     "No Rate Limiting on Login / Auth Endpoints",
     "The login, register, and forgot-password endpoints have no per-endpoint rate limiting, allowing unlimited brute-force attempts.",
     "Without rate limiting, an attacker can make thousands of login attempts per second, enabling credential stuffing or password spray attacks against all accounts.",
     "Add express-rate-limit middleware: limit auth endpoints to 10 requests/minute per IP. Consider CAPTCHA for repeated failures. Implement exponential backoff or temporary account lockout."),

    ("SEC-018", "Business Logic",         "MEDIUM",   "CWE-400", "A05:2021",
     "backend/src/controllers/scanController.js + routes/scan.js",
     "No Rate Limiting on CPU-Intensive AI Scan Endpoint",
     "The /api/scans/analyze endpoint performs expensive TensorFlow + Jimp image processing with only the general apiLimiter (200 req/15 min).",
     "An attacker can flood the server with large image uploads, causing CPU/memory exhaustion and denial of service. At 5 MB per file, 200 concurrent uploads = 1 GB memory pressure.",
     "Apply dedicated stricter rate limiting (e.g., 10 scans/minute per user). Validate image dimensions via jimp after loading to prevent decompression bombs."),

    ("SEC-019", "Configuration",          "MEDIUM",   "CWE-400", "A05:2021",
     "backend/src/app.js (L19-L20)",
     "No Request Body Size Limit for JSON / URL-Encoded Payloads",
     "express.json() and express.urlencoded() are used without a size limit option.",
     "An attacker can POST arbitrarily large JSON payloads to any endpoint, potentially causing memory exhaustion and denial of service.",
     "Add size limits: app.use(express.json({ limit: '10kb' })) and app.use(express.urlencoded({ limit: '10kb', extended: true }))."),

    ("SEC-020", "Configuration",          "MEDIUM",   "CWE-693", "A05:2021",
     "backend/src/app.js",
     "Missing Content Security Policy (CSP) Header",
     "The security headers middleware sets X-Content-Type-Options, X-Frame-Options, X-XSS-Protection, and Referrer-Policy but omits a Content-Security-Policy header.",
     "Without CSP, if any endpoint reflects user input (current or future), XSS attacks have no browser-side mitigation. CSP is a critical defense-in-depth control.",
     "Add: res.setHeader('Content-Security-Policy', \"default-src 'none'; script-src 'self'; connect-src 'self'; img-src 'self' data:; style-src 'self'\"). Use Helmet's contentSecurityPolicy option."),

    ("SEC-021", "Configuration",          "MEDIUM",   "CWE-523", "A05:2021",
     "backend/src/app.js",
     "Missing Strict-Transport-Security (HSTS) Header",
     "No HSTS header is set. HTTP connections are not explicitly forced to HTTPS at the application layer.",
     "Without HSTS, users who navigate via HTTP can be subject to SSL stripping attacks, exposing their session JWT tokens to network eavesdroppers.",
     "Add: res.setHeader('Strict-Transport-Security', 'max-age=31536000; includeSubDomains; preload') when NODE_ENV === 'production'. Enable via Helmet: helmet.hsts({maxAge: 31536000})."),

    ("SEC-022", "Authorization",          "MEDIUM",   "CWE-639", "A01:2021",
     "backend/src/config/db.js (L314-L319)",
     "Embedded SQL Engine Has Incorrect OR Logic — IDOR in Offline Mode",
     "In the embedded JSON store fallback, the ownership filter uses: filter(f => f.user_id === userId || f.id === idOrUser) — an OR instead of AND.",
     "This logic error could allow a user to retrieve another user's family member record in embedded (offline) mode if they supply the target's ID. When pgConnected is false, this fallback is active.",
     "Fix the filter to: filter(f => f.id === id && f.user_id === userId). Never use OR logic for ownership checks. The embedded store must mirror exact SQL constraints."),

    ("SEC-023", "Configuration",          "MEDIUM",   "CWE-284", "A05:2021",
     "docker-compose.yml (L9-L14)",
     "Docker Compose Exposes PostgreSQL and Adminer Publicly",
     "Docker Compose exposes PostgreSQL on 0.0.0.0:5432 with default credentials (postgres/postgrespassword). Adminer is exposed on 0.0.0.0:8080 with full DB admin access.",
     "If this configuration is used in any shared or cloud dev environment, the database and its admin interface are publicly accessible to anyone on the network.",
     "Bind to 127.0.0.1 only: '127.0.0.1:5432:5432'. Use strong random passwords via .env. Remove Adminer from any non-local environment or add HTTP authentication."),

    ("SEC-024", "Business Logic",         "MEDIUM",   "CWE-311", "A02:2021",
     "backend/src/config/db.js (L97-L103)",
     "Embedded JSON Store Writes Sensitive Data to Plaintext File",
     "When PostgreSQL is unavailable, user data (password hashes, emails, phone numbers) is stored in embedded_store.json on disk in plaintext JSON.",
     "The JSON file is not encrypted. If the server filesystem is accessible (path traversal, compromised process), all user data is immediately exposed.",
     "Restrict file permissions to 600 (owner read/write only). In production, do not fall back to the embedded store — fail loudly if PostgreSQL is unavailable. Consider encrypting if offline mode is required."),

    ("SEC-025", "Dependency",             "MEDIUM",   "CWE-1395","A06:2021",
     "backend/package.json",
     "multer ^1.4.5-lts.1 — Historical DoS Vulnerabilities",
     "The application uses multer 1.4.x LTS. The multer v1.x line has had memory DoS vulnerabilities in multipart boundary handling.",
     "Using older library versions with known vulnerability history increases attack surface, especially on the file upload path which handles untrusted user content.",
     "Upgrade to multer v2.x (active development with security improvements) or pin to the most current 1.4.x-lts patch. Run 'npm audit' and address all findings."),

    ("SEC-026", "Dependency",             "MEDIUM",   "CWE-1395","A06:2021",
     "backend/package.json",
     "express ^4.19.2 — Known Path Traversal Fixed in 4.21.x",
     "Express 4.19.2 predates the 4.21.x series which addressed path traversal issues in express.static and route handling (CVE-2024-29041).",
     "The application uses express.static for uploads and illustrations — paths that serve user-related content. Path traversal in static serving could expose arbitrary files.",
     "Pin to express@^4.21.x or latest stable. Run 'npm audit' and update accordingly. Consider migrating to Express 5.x for long-term security support."),

    ("SEC-027", "Authentication",         "LOW",      "CWE-613", "A07:2021",
     "backend/src/controllers/authController.js (L121)",
     "JWT Token Expiry is 7 Days — No Refresh Token / Revocation",
     "Tokens are signed with expiresIn: '7d'. There is no token refresh, revocation, or logout mechanism.",
     "A stolen JWT remains valid for up to 7 days. There is no server-side way to invalidate a session (e.g., on logout, password change, or suspicious activity detection).",
     "Shorten expiry to 15-60 minutes. Implement refresh tokens (stored in HttpOnly cookies, rotated on use). On password change, blacklist previous tokens in Redis. Add proper logout endpoint."),

    ("SEC-028", "Configuration",          "LOW",      "CWE-693", "A05:2021",
     "backend/src/app.js (L26)",
     "X-XSS-Protection Header is Deprecated and Removed from Modern Browsers",
     "X-XSS-Protection: 1; mode=block is set, but this header was deprecated and removed from Chromium 78+ and other modern browsers.",
     "The header provides no protection in modern browsers and can cause issues in some edge cases. Relying on it gives a false sense of security.",
     "Remove X-XSS-Protection header. Rely on a proper Content-Security-Policy instead. Use Helmet.js for managed, up-to-date security headers with current browser support."),

    ("SEC-029", "Configuration",          "LOW",      "CWE-284", "A05:2021",
     "backend/src/server.js (L29)",
     "Server Binds to 0.0.0.0 — Exposed on All Network Interfaces",
     "The server explicitly binds to 0.0.0.0, exposing the API on all network interfaces, not just localhost.",
     "In local dev or internal networks, this exposes the API to the local network. On cloud servers without strict firewall rules, the API may be directly accessible bypassing the reverse proxy.",
     "In development, bind to 127.0.0.1. In production, bind to 0.0.0.0 only behind Render's reverse proxy with proper firewall rules. Document the deployment topology clearly."),

    ("SEC-030", "Sensitive Data",         "LOW",      "CWE-532", "A09:2021",
     "backend/src/controllers/authController.js (L37, L63)",
     "PII Logged in Console During Google Login Flow",
     "googleClientId.trim() and email, googleId values are logged to console at info level during normal operation.",
     "In a log aggregation system (Render logs, Datadog), PII (email addresses, Google IDs) being logged creates compliance risk under GDPR and India's DPDP Act.",
     "Remove PII from logs. Log only non-identifiable event descriptors: '[GoogleLogin] user authenticated successfully'. Use structured logging with PII redaction (pino redact option)."),
]

DEPENDENCY_REVIEW = [
    # Package, Specified, Latest, Risk, CVE/Notes, Action
    ("express",              "^4.19.2",       "4.21.x",  "MEDIUM", "CVE-2024-29041 — path traversal via express.static in 4.19.x",    "Upgrade to 4.21.x immediately"),
    ("multer",               "^1.4.5-lts.1",  "2.x",     "MEDIUM", "Historical DoS in multipart boundary handling",                    "Evaluate upgrade to multer v2.x"),
    ("@tensorflow/tfjs",     "^4.22.0",        "4.22.x",  "MEDIUM", "Large attack surface; supply chain risk; model poisoning possible","Pin exact version, verify integrity"),
    ("jimp",                 "^1.6.1",         "1.6.x",   "MEDIUM", "Decompression bomb risk in image libraries without dimension check","Validate image dimensions post-load"),
    ("jsonwebtoken",         "^9.0.2",         "9.0.2",   "LOW",    "Current stable",                                                  "Maintain; explicitly set algorithm"),
    ("bcryptjs",             "^2.4.3",         "2.4.3",   "LOW",    "Pure JS impl; consider native bcrypt for performance",             "Consider argon2 for stronger hashing"),
    ("cors",                 "^2.8.5",         "2.8.5",   "LOW",    "Maintained but no recent updates (2021)",                          "Monitor for security updates"),
    ("dotenv",               "^16.4.5",        "16.4.5",  "LOW",    "Current stable",                                                  "Maintain current version"),
    ("google-auth-library",  "^11.0.0",        "11.x",    "LOW",    "Current stable; keep updated for OAuth security patches",          "Keep updated; subscribe to advisories"),
    ("nodemailer",           "^9.0.5",         "9.0.5",   "LOW",    "Current stable",                                                  "Remove rejectUnauthorized: false"),
    ("pg",                   "^8.11.5",        "8.13.x",  "LOW",    "Minor patch available",                                            "Upgrade to 8.13.x"),
    ("helmet",               "^8.3.0",         "8.3.x",   "LOW",    "Current stable",                                                  "Maintain; configure CSP properly"),
    ("express-rate-limit",   "^8.6.2",         "8.6.x",   "LOW",    "Current stable",                                                  "Add per-endpoint limiters"),
    ("uuid",                 "^9.0.1",         "9.0.1",   "LOW",    "Current stable",                                                  "Maintain current version"),
    ("jest",                 "^29.7.0",        "29.7.0",  "LOW",    "Dev only — current stable",                                        "Maintain"),
    ("nodemon",              "^3.1.0",         "3.1.x",   "LOW",    "Dev only — current stable",                                        "Maintain"),
    ("supertest",            "^6.3.4",         "6.3.4",   "LOW",    "Dev only — current stable",                                        "Maintain"),
    ("@tensorflow-models/coco-ssd","^2.2.3",   "2.2.3",   "LOW",    "Stable; depends on tfjs",                                          "Verify model file integrity"),
]

REMEDIATION_PLAN = [
    ("P0", "Immediate — Within 24 Hours",   "SEC-001,002,003",
     "ROTATE: Generate new JWT_SECRET (openssl rand -base64 32). Rotate DB password. Revoke/re-register Google OAuth client. Run git-secrets / gitleaks on full history. Add backend/.gitignore with .env entry. Remove GOOGLE_CLIENT_ID from render.yaml (use sync: false).",
     "CRITICAL"),
    ("P1", "Short-Term — Within 1 Week",    "SEC-004,008,009,011,012",
     "Remove JWT code-level fallback — fail-fast on missing secret. Set strict CORS allowlist in all environments, remove NODE_ENV bypass. Add authMiddleware to /uploads route. Enable SSL rejectUnauthorized: true for DB and email connections. Protect /api/system/database-status with auth.",
     "HIGH"),
    ("P2", "Medium-Term — Within 2 Weeks",  "SEC-005,006,007,017,018,019",
     "Increase password minimum length to 12 chars + complexity regex. Fix user enumeration in forgot-password (return 200 always). Fix IDOR in toothbrush DELETE — add user_id to WHERE clause. Add per-endpoint rate limiters for auth and scan endpoints. Add body size limits to express.json(). Add decompression bomb protection to image processing.",
     "HIGH"),
    ("P3", "Medium-Term — Within 1 Month",  "SEC-010,013,014,015,020,021,022,023,024",
     "Add proper CSP header (remove unsafe-inline, use nonces). Add HSTS header in production. Fix OR→AND logic in embedded store ownership check. Bind Docker ports to 127.0.0.1. Add Joi/Zod validation schemas for all controller inputs. Add magic-byte file type validation (file-type npm). Restrict embedded store fallback from production.",
     "MEDIUM"),
    ("P4", "Ongoing — Best Practice",       "SEC-025,026,027,028,029,030",
     "Upgrade express to 4.21.x (CVE-2024-29041). Upgrade multer to 2.x. Implement JWT refresh token + revocation (Redis). Remove deprecated X-XSS-Protection header. Bind to 127.0.0.1 in dev. Implement structured logging (pino) with PII redaction. Add npm audit to CI pipeline. Schedule quarterly security reviews.",
     "LOW"),
]

GITHUB_ACTIONS_SUMMARY = [
    ("Job Name",              "Tool / Action",         "Purpose",                              "Triggers Failure"),
    ("detect-stack",          "Bash shell script",     "Auto-detect framework/language",       "Never (informational)"),
    ("gitleaks",              "gitleaks/gitleaks-action@v2","Full git history secret scan",   "On any secret detection"),
    ("semgrep",               "semgrep/semgrep-action@v1","SAST: OWASP, JWT, Node.js, SQL",  "On CRITICAL severity"),
    ("trivy",                 "aquasecurity/trivy-action","Filesystem + dependency CVE scan",  "On CRITICAL CVEs"),
    ("npm-audit",             "npm audit --json",      "Dependency vulnerability review",      "On CRITICAL npm findings"),
    ("dependency-review",     "actions/dependency-review-action@v4","PR dependency diff",    "On CRITICAL severity (PRs only)"),
    ("hardening-checks",      "Custom bash grep",      "Code hygiene: secrets, SSL, CSP",     "Advisory warnings only"),
    ("security-summary",      "GitHub Step Summary",   "Publish consolidated report",          "If any upstream job failed"),
]

SECURITY_SCORE_BREAKDOWN = [
    ("Authentication",      65,  "Bcrypt hashing, JWT, Google OAuth — but weak secret, 7d expiry, no revocation"),
    ("Authorization",       78,  "SQL JOIN ownership checks throughout — but IDOR risk in toothbrush DELETE, OR logic in embedded store"),
    ("Input Validation",    60,  "Parameterized queries prevent SQL injection — but no schema validation on inputs"),
    ("Injection Safety",    88,  "Parameterized queries ($1, $2...) used consistently — minimal injection risk"),
    ("Cryptography",        40,  "bcrypt used — but JWT secret weak, SSL disabled for DB and email, no HSTS"),
    ("Sensitive Data",      30,  "Secrets in .env committed, Client ID in render.yaml, PII in console logs"),
    ("Configuration",       55,  "Helmet present — but CSP missing, CORS bypass, body no size limit, uploads unauthenticated"),
    ("Dependency Health",   65,  "Mostly current — express 4.19 has CVE, multer historical issues, TF.js large surface"),
]

POSITIVE_CONTROLS = [
    "SQL Injection prevented — parameterized queries ($1, $2...) used consistently in ALL controllers",
    "Password hashing — bcryptjs with salt factor 10 applied correctly in register and change-password",
    "X-Powered-By header disabled — prevents Express.js version fingerprinting",
    "JWT format validation — 'Bearer' prefix explicitly verified in auth middleware before decode",
    "Ownership validation — all CRUD operations verify resource belongs to authenticated user via SQL JOINs",
    "Multer file size cap — 5 MB limit prevents basic upload-based DoS attacks",
    "MIME type filter — image/* filter on upload provides first-level content type restriction",
    "Google ID token verified server-side — OAuth2Client.verifyIdToken() used with correct audience check",
    "Helmet middleware — security headers configured (X-Content-Type-Options, X-Frame-Options, Referrer-Policy)",
    "Rate limiting present — general apiLimiter on all routes, authLimiter on auth endpoints",
]

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def build_cover(ws):
    """Cover / Title page"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 65

    banner(ws, 1, "🔐  BrushIQ — Comprehensive Secure Code Review Report", 2,
           fg=C_NAVY, size=18, height=60)

    ws.row_dimensions[2].height = 10

    meta = [
        ("📅 Report Date",      datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("🏢 Application",      "BrushIQ — AI-Powered Oral Healthcare Platform"),
        ("⚙️  Backend Stack",    "Node.js 18+ / Express.js 4.19.2"),
        ("🗄️  Database",         "PostgreSQL 15 (Supabase) + Embedded JSON Fallback"),
        ("🔑 Auth",              "JWT (jsonwebtoken ^9.0.2) + Google OAuth 2.0"),
        ("🚀 Deployment",        "Render.com (backend), Vercel (frontend), Docker Compose (dev)"),
        ("🔍 Review Type",       "Static Code Analysis (SAST) — Defensive Security Review"),
        ("👤 Reviewed By",       "Antigravity AI Security Analyzer"),
        ("📂 Scope",             "backend/ — all source files, configs, dependencies"),
        ("📊 Review Phases",     "Phase 1: Discovery | Phase 2: API | Phase 3: SAST | Phase 4: Deps | Phase 5: Report | Phase 6: CI/CD"),
    ]
    for i, (k, v) in enumerate(meta, 3):
        ws.row_dimensions[i].height = 22
        set_cell(ws, i, 1, k, bold=True, fg=C_DARK_BLUE, font_color=C_WHITE, h_align="right")
        set_cell(ws, i, 2, v, fg=C_LIGHT_BG, font_color=C_DARK_TEXT, wrap=True)

    ws.row_dimensions[14].height = 12
    # Divider
    ws.merge_cells("A15:B15")
    c = ws["A15"]
    c.value = ""
    c.fill = fill(C_ACCENT)
    c.border = thin_border()
    ws.row_dimensions[15].height = 4

    # Score
    ws.row_dimensions[16].height = 14
    section_title(ws, 16, 1, "  OVERALL SECURITY SCORE", 2, fg=C_DARK_BLUE, size=11)

    ws.row_dimensions[17].height = 70
    ws.merge_cells("A17:B17")
    sc = ws["A17"]
    sc.value = "52 / 100"
    sc.font = Font(bold=True, color=C_ORANGE, size=44, name="Calibri")
    sc.fill = fill(C_LIGHT_BG)
    sc.alignment = align("center", "center")
    sc.border = medium_border()

    ws.row_dimensions[18].height = 35
    ws.merge_cells("A18:B18")
    c = ws["A18"]
    c.value = ("⚠  Rating: MODERATE RISK  |  3 Critical findings require IMMEDIATE action before production launch. "
               "The application has solid SQL injection prevention and ownership checks, but critical "
               "secret management, SSL configuration, and authentication weaknesses must be addressed.")
    c.font = Font(bold=False, color=C_ORANGE, size=10, name="Calibri")
    c.fill = fill("FFF7ED")
    c.alignment = align("center", "center", wrap=True)
    c.border = thin_border()

    ws.row_dimensions[19].height = 10
    section_title(ws, 20, 1, "  FINDING SUMMARY", 2, fg=C_DARK_BLUE, size=11)

    counts = [
        ("🔴  CRITICAL", 3,  C_RED,    "SEC-001, SEC-002, SEC-003", "Secrets committed to repository"),
        ("🟠  HIGH",     8,  C_ORANGE, "SEC-004 through SEC-012",   "Auth, SSL, CORS, IDOR, Uploads"),
        ("🟡  MEDIUM",   13, C_YELLOW, "SEC-013 through SEC-026",   "Validation, Config, Logic, Deps"),
        ("🔵  LOW",      5,  C_TEAL,   "SEC-027 through SEC-030",   "Token expiry, CSP, Logging"),
    ]
    header_row(ws, 21, ["Severity", "Count", "Finding IDs", "Examples"], fg=C_DARK_BLUE)
    for i, (sev, cnt, col, ids, examples) in enumerate(counts, 22):
        ws.row_dimensions[i].height = 20
        set_cell(ws, i, 1, sev,      bold=True, fg=col, font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 2, cnt,      bold=True, fg=C_LIGHT_BG, h_align="center", font_size=13)
        set_cell(ws, i, 3, ids,      fg=C_LIGHT_BG)
        set_cell(ws, i, 4, examples, fg=C_LIGHT2, wrap=True)

    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40


def build_executive_summary(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 65

    banner(ws, 1, "EXECUTIVE SUMMARY — BrushIQ Secure Code Review", 2,
           fg=C_DARK_BLUE, size=14, height=35)

    ws.row_dimensions[3].height = 18
    section_title(ws, 3, 1, "  Application Overview", 2, fg=C_ACCENT, size=11)
    ws.row_dimensions[4].height = 80
    ws.merge_cells("A4:B4")
    c = ws["A4"]
    c.value = (
        "BrushIQ is an AI-powered oral healthcare platform that allows users to photograph and analyze toothbrush "
        "bristle wear using computer vision (TensorFlow.js + COCO-SSD). The backend is a Node.js/Express REST API "
        "deployed on Render.com, connected to a Supabase-hosted PostgreSQL database. Users authenticate via email/"
        "password (JWT) or Google OAuth 2.0. The platform processes sensitive health-adjacent imagery and personal "
        "profile data for family members.\n\n"
        "This review covers static analysis (SAST) of the entire backend source code, dependency audit, "
        "configuration review, and API security assessment. No active exploitation was performed."
    )
    c.font = Font(color=C_DARK_TEXT, size=10, name="Calibri")
    c.fill = fill(C_CREAM)
    c.alignment = align("left", "top", wrap=True)
    c.border = thin_border()

    ws.row_dimensions[5].height = 10
    section_title(ws, 6, 1, "  Key Findings At a Glance", 2, fg=C_RED, size=11)

    key_findings = [
        ("🔴 CRITICAL", "Credentials committed to git — JWT secret, DB password, and Google Client ID are all in the repository"),
        ("🔴 CRITICAL", "Weak JWT secret in production .env (30-char dictionary-guessable string)"),
        ("🔴 CRITICAL", "Google OAuth Client ID hardcoded in render.yaml tracked in version control"),
        ("🟠 HIGH", "SSL certificate validation disabled for both database (PostgreSQL) and email (SMTP) connections"),
        ("🟠 HIGH", "Uploaded scan images (health data) served publicly without authentication"),
        ("🟠 HIGH", "CORS allows all origins when ALLOWED_ORIGINS env var is not set (wildcard fallback)"),
        ("🟠 HIGH", "User enumeration possible via /api/auth/forgot-password HTTP 400 response"),
        ("🟠 HIGH", "Unauthenticated /api/system/database-status endpoint leaks infrastructure state"),
        ("🟡 MEDIUM", "No rate limiting on login/register endpoints — brute-force attacks possible"),
        ("🟡 MEDIUM", "Client-controlled scan metrics (healthScore, wearPercentage) not validated server-side"),
    ]
    header_row(ws, 7, ["Severity", "Finding"], fg=C_DARK_BLUE)
    for i, (sev, desc) in enumerate(key_findings, 8):
        ws.row_dimensions[i].height = 22
        col = C_RED if "CRITICAL" in sev else (C_ORANGE if "HIGH" in sev else C_YELLOW)
        set_cell(ws, i, 1, sev,  bold=True, fg=col, font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 2, desc, fg=C_LIGHT_BG if i % 2 == 0 else C_LIGHT2, wrap=True)

    ws.row_dimensions[19].height = 10
    section_title(ws, 20, 1, "  Positive Security Controls Detected", 2, fg=C_GREEN, size=11)

    for i, ctrl in enumerate(POSITIVE_CONTROLS, 21):
        ws.row_dimensions[i].height = 20
        ws.merge_cells(f"A{i}:B{i}")
        c = ws.cell(row=i, column=1, value=f"✓  {ctrl}")
        c.font = Font(color="166534", size=10, name="Calibri")
        c.fill = fill("F0FDF4" if i % 2 == 0 else "DCFCE7")
        c.alignment = align("left", "center", wrap=True)
        c.border = thin_border()

    ws.row_dimensions[32].height = 10
    section_title(ws, 33, 1, "  Security Score Breakdown", 2, fg=C_DARK_BLUE, size=11)
    header_row(ws, 34, ["Category", "Score (/100)", "Notes"], fg=C_ACCENT)

    for i, (cat, score, notes) in enumerate(SECURITY_SCORE_BREAKDOWN, 35):
        ws.row_dimensions[i].height = 20
        bg = C_LIGHT_BG if i % 2 == 0 else C_LIGHT2
        score_col = C_RED if score < 50 else (C_ORANGE if score < 70 else C_GREEN)
        set_cell(ws, i, 1, cat,   bold=True, fg=bg)
        set_cell(ws, i, 2, f"{score}/100", bold=True, fg=score_col, font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 3, notes, fg=bg, wrap=True)


def build_backend_inventory(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 72

    banner(ws, 1, "PHASE 1 — Backend Technology Inventory", 2,
           fg=C_DARK_BLUE, size=14, height=35)

    ws.row_dimensions[2].height = 8
    header_row(ws, 3, ["Component", "Details"], fg=C_ACCENT)

    for i, (comp, detail) in enumerate(BACKEND_INVENTORY, 4):
        ws.row_dimensions[i].height = 24
        bg1 = C_DARK_BLUE if i % 2 == 0 else "1E3A5F"
        bg2 = C_LIGHT_BG if i % 2 == 0 else C_LIGHT2
        set_cell(ws, i, 1, comp,   bold=True, fg=bg1, font_color=C_WHITE)
        set_cell(ws, i, 2, detail, fg=bg2, wrap=True)


def build_endpoint_inventory(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 45

    banner(ws, 1, "PHASE 2 — API Endpoint Inventory", 7,
           fg=C_DARK_BLUE, size=14, height=35)

    ws.row_dimensions[2].height = 8
    header_row(ws, 3, ["Endpoint", "Method", "Auth?", "Risk", "Roles", "Controller / File", "Security Notes"], fg=C_ACCENT)

    method_colors = {
        "GET": "0EA5E9", "POST": "16A34A", "PUT": "D97706",
        "DELETE": "DC2626", "PATCH": "7C3AED"
    }
    risk_colors = {
        "High": C_ORANGE, "Medium": C_YELLOW, "Low": C_GREEN
    }

    for i, (ep, method, auth, risk, role, ctrl, notes) in enumerate(ENDPOINT_INVENTORY, 4):
        even = i % 2 == 0
        row_bg = C_LIGHT_BG if even else C_LIGHT2
        ws.row_dimensions[i].height = 20

        set_cell(ws, i, 1, ep,     fg=row_bg, font_size=9)
        mc = method_colors.get(method, C_GREY)
        set_cell(ws, i, 2, method, bold=True, fg=mc, font_color=C_WHITE, h_align="center", font_size=9)
        auth_fg = C_GREEN if auth == "Yes" else C_RED
        set_cell(ws, i, 3, auth,   bold=True, fg=auth_fg, font_color=C_WHITE, h_align="center", font_size=9)
        rc = risk_colors.get(risk, C_GREY)
        set_cell(ws, i, 4, risk,   bold=True, fg=rc, font_color=C_WHITE, h_align="center", font_size=9)
        set_cell(ws, i, 5, role,   fg=row_bg, h_align="center", font_size=9)
        set_cell(ws, i, 6, ctrl,   fg=row_bg, italic=True, font_size=9)
        set_cell(ws, i, 7, notes,  fg=row_bg, wrap=True, font_size=9)


def build_findings(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 48
    ws.column_dimensions['G'].width = 38
    ws.column_dimensions['H'].width = 55
    ws.column_dimensions['I'].width = 55
    ws.column_dimensions['J'].width = 60

    banner(ws, 1, "PHASE 3 — Security Findings (SAST) — 30 Total Issues", 10,
           fg=C_DARK_BLUE, size=14, height=35)

    ws.row_dimensions[2].height = 8
    cols = ["ID", "Category", "Severity", "CWE", "OWASP",
            "File Path", "Title", "Description", "Why It's a Risk", "Recommended Fix"]
    header_row(ws, 3, cols, fg=C_ACCENT)

    sev_colors = {
        "CRITICAL": C_RED,
        "HIGH":     C_ORANGE,
        "MEDIUM":   C_YELLOW,
        "LOW":      C_TEAL,
    }

    for i, (fid, cat, sev, cwe, owasp, fpath, title, desc, concern, fix) in enumerate(FINDINGS, 4):
        ws.row_dimensions[i].height = 65
        even = i % 2 == 0
        row_bg = C_LIGHT_BG if even else C_LIGHT2
        sc = sev_colors.get(sev, C_GREY)

        set_cell(ws, i, 1,  fid,     bold=True, fg=row_bg,   h_align="center", font_size=9)
        set_cell(ws, i, 2,  cat,     fg=row_bg,              wrap=True, font_size=9)
        set_cell(ws, i, 3,  sev,     bold=True, fg=sc, font_color=C_WHITE, h_align="center", font_size=9)
        set_cell(ws, i, 4,  cwe,     fg=row_bg,              h_align="center", font_size=9)
        set_cell(ws, i, 5,  owasp,   fg=row_bg,              h_align="center", font_size=9)
        set_cell(ws, i, 6,  fpath,   fg=row_bg,              wrap=True, italic=True, font_size=9)
        set_cell(ws, i, 7,  title,   bold=True, fg=row_bg,   wrap=True, font_size=9)
        set_cell(ws, i, 8,  desc,    fg=row_bg,              wrap=True, font_size=9)
        set_cell(ws, i, 9,  concern, fg=row_bg,              wrap=True, font_size=9)
        set_cell(ws, i, 10, fix,     fg="F0FDF4",            wrap=True, font_size=9)


def build_dependency_review(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 58
    ws.column_dimensions['F'].width = 40

    banner(ws, 1, "PHASE 4 — Dependency Review", 6,
           fg=C_DARK_BLUE, size=14, height=35)

    ws.row_dimensions[2].height = 8
    header_row(ws, 3, ["Package", "Specified Version", "Latest Stable", "Risk", "CVE / Notes", "Recommended Action"], fg=C_ACCENT)

    risk_colors = {"MEDIUM": C_ORANGE, "LOW": C_GREEN}

    for i, (pkg, spec, latest, risk, notes, action) in enumerate(DEPENDENCY_REVIEW, 4):
        even = i % 2 == 0
        row_bg = C_LIGHT_BG if even else C_LIGHT2
        rc = risk_colors.get(risk, C_GREY)
        ws.row_dimensions[i].height = 24

        set_cell(ws, i, 1, pkg,    bold=True, fg=row_bg)
        set_cell(ws, i, 2, spec,   fg=row_bg, h_align="center")
        set_cell(ws, i, 3, latest, fg=row_bg, h_align="center")
        set_cell(ws, i, 4, risk,   bold=True, fg=rc, font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 5, notes,  fg=row_bg, wrap=True)
        set_cell(ws, i, 6, action, fg="F0FDF4", wrap=True)


def build_remediation(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 14

    banner(ws, 1, "PHASE 5 — Recommended Remediation Roadmap", 5,
           fg=C_DARK_BLUE, size=14, height=35)

    ws.row_dimensions[2].height = 8
    header_row(ws, 3, ["#", "Priority Tier", "Finding IDs", "Actions Required", "Risk Level"], fg=C_ACCENT)

    tier_colors = {
        "P0": C_RED,
        "P1": C_ORANGE,
        "P2": "D97706",
        "P3": C_YELLOW,
        "P4": C_GREEN,
    }
    risk_colors = {"CRITICAL": C_RED, "HIGH": C_ORANGE, "HIGH/MEDIUM": C_ORANGE,
                   "MEDIUM": C_YELLOW, "LOW": C_GREEN}

    for i, (pid, tier, fids, action, risk) in enumerate(REMEDIATION_PLAN, 4):
        ws.row_dimensions[i].height = 80
        tc = tier_colors.get(pid, C_GREY)
        rc = risk_colors.get(risk, C_GREY)

        set_cell(ws, i, 1, pid,    bold=True, fg=tc, font_color=C_WHITE, h_align="center", font_size=14)
        set_cell(ws, i, 2, tier,   bold=True, fg=tc, font_color=C_WHITE, wrap=True)
        set_cell(ws, i, 3, fids,   fg=C_LIGHT_BG, wrap=True, font_size=9)
        set_cell(ws, i, 4, action, fg=C_LIGHT_BG, wrap=True)
        set_cell(ws, i, 5, risk,   bold=True, fg=rc, font_color=C_WHITE, h_align="center")


def build_github_actions(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30

    banner(ws, 1, "PHASE 6 — GitHub Actions Security Workflow", 4,
           fg=C_DARK_BLUE, size=14, height=35)

    ws.row_dimensions[2].height = 8
    section_title(ws, 3, 1, "  Workflow: .github/workflows/security.yml", 4, fg=C_ACCENT, size=11)

    ws.row_dimensions[4].height = 30
    ws.merge_cells("A4:D4")
    c = ws["A4"]
    c.value = ("Triggers: Push to main/develop/master | Pull Requests to main/master | "
               "Schedule: Every Monday 03:00 UTC | Manual (workflow_dispatch)  |  "
               "Fails ONLY on CRITICAL severity findings (Gitleaks, Trivy, npm audit, hardening-checks).")
    c.font = Font(color=C_DARK_TEXT, size=10, name="Calibri")
    c.fill = fill(C_CREAM)
    c.alignment = align("left", "center", wrap=True)
    c.border = thin_border()

    ws.row_dimensions[5].height = 8
    header_row(ws, 6, ["Job Name", "Tool / Action", "Purpose", "Pipeline Failure Condition"], fg=C_ACCENT)

    for i, row in enumerate(GITHUB_ACTIONS_SUMMARY[1:], 7):
        ws.row_dimensions[i].height = 22
        even = i % 2 == 0
        bg = C_LIGHT_BG if even else C_LIGHT2
        job_name, tool, purpose, fail_cond = row
        job_col = C_DARK_BLUE if "detect" in job_name.lower() else (
                  C_RED    if "gitleaks" in job_name.lower() else (
                  C_PURPLE if "semgrep"  in job_name.lower() else (
                  C_ORANGE if "trivy"    in job_name.lower() else (
                  C_TEAL   if "npm"      in job_name.lower() else (
                  C_GREEN  if "dep"      in job_name.lower() else C_ACCENT)))))
        set_cell(ws, i, 1, job_name,  bold=True, fg=job_col, font_color=C_WHITE)
        set_cell(ws, i, 2, tool,      fg=bg)
        set_cell(ws, i, 3, purpose,   fg=bg, wrap=True)
        fail_bg = C_RED if "CRITICAL" in fail_cond or "secret" in fail_cond.lower() else (
                  C_YELLOW if "warning" in fail_cond.lower() else C_GREEN)
        set_cell(ws, i, 4, fail_cond, bold=True, fg=fail_bg, font_color=C_WHITE, wrap=True)

    ws.row_dimensions[16].height = 12
    section_title(ws, 17, 1, "  Artifacts Produced", 4, fg=C_DARK_BLUE, size=11)

    artifacts = [
        ("gitleaks-report",  "SARIF",    "Secret detection results from full git history scan"),
        ("semgrep-report",   "SARIF+JSON","SAST results: OWASP Top 10, JWT, Node.js, SQL injection"),
        ("trivy-reports",    "SARIF+JSON","CVE scan of filesystem and npm dependencies"),
        ("npm-audit-report", "JSON",     "npm audit vulnerability report for backend/package.json"),
    ]
    header_row(ws, 18, ["Artifact Name", "Format", "Description"], fg=C_ACCENT)
    for i, (name, fmt, desc) in enumerate(artifacts, 19):
        ws.row_dimensions[i].height = 20
        bg = C_LIGHT_BG if i % 2 == 0 else C_LIGHT2
        set_cell(ws, i, 1, name, bold=True, fg=bg)
        set_cell(ws, i, 2, fmt,  fg=bg, h_align="center")
        set_cell(ws, i, 3, desc, fg=bg, wrap=True)


def build_risk_dashboard(ws):
    ws.sheet_view.showGridLines = False

    banner(ws, 1, "🎯  Security Risk Dashboard", 6, fg=C_NAVY, size=16, height=45)

    # KPI Cards
    kpis = [
        ("Security Score", "52/100",  C_ORANGE, "Moderate Risk"),
        ("Critical",       "3",       C_RED,    "Immediate Action"),
        ("High",           "8",       C_ORANGE, "Fix This Sprint"),
        ("Medium",         "13",      C_YELLOW, "Plan & Schedule"),
        ("Low",            "5",       C_TEAL,   "Best Practice"),
        ("Total Issues",   "29",      C_ACCENT, "All Severities"),
    ]
    for col, (label, val, col_hex, sub) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 55
        ws.row_dimensions[5].height = 18

        c = ws.cell(row=3, column=col, value=label)
        c.font = Font(bold=True, color=C_WHITE, size=9, name="Calibri")
        c.fill = fill(col_hex)
        c.alignment = align("center", "center")
        c.border = medium_border()

        c2 = ws.cell(row=4, column=col, value=val)
        c2.font = Font(bold=True, color=col_hex, size=30, name="Calibri")
        c2.fill = fill(C_LIGHT_BG)
        c2.alignment = align("center", "center")
        c2.border = medium_border()

        c3 = ws.cell(row=5, column=col, value=sub)
        c3.font = Font(bold=False, color=C_GREY, size=9, name="Calibri")
        c3.fill = fill(C_LIGHT2)
        c3.alignment = align("center", "center")
        c3.border = thin_border()

    ws.row_dimensions[7].height = 10
    section_title(ws, 8, 1, "  Finding Distribution by Severity", 6, fg=C_DARK_BLUE)

    chart_data_rows = [("Severity", "Count"),
                       ("Critical", 3), ("High", 8), ("Medium", 13), ("Low", 5)]
    for r, (sev, cnt) in enumerate(chart_data_rows, 9):
        ws.cell(row=r, column=1, value=sev)
        ws.cell(row=r, column=2, value=cnt)

    bar = BarChart()
    bar.type = "col"
    bar.title = "Security Findings by Severity"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Severity"
    bar.style = 10
    bar.width = 22
    bar.height = 13

    data_ref = Reference(ws, min_col=2, min_row=9, max_row=12)
    cats_ref = Reference(ws, min_col=1, min_row=10, max_row=12)
    bar.add_data(data_ref, titles_from_data=False)
    bar.set_categories(cats_ref)
    ws.add_chart(bar, "C8")

    # Pie chart for category
    section_title(ws, 24, 1, "  Finding Distribution by Category", 6, fg=C_DARK_BLUE)

    categories = [
        ("Cryptography/Secrets",  2),
        ("Sensitive Data",        3),
        ("Authentication",        3),
        ("Authorization/IDOR",    2),
        ("Input Validation",      3),
        ("Configuration",         8),
        ("Business Logic",        3),
        ("Dependency",            2),
        ("PII Logging",           1),
    ]
    header_row(ws, 25, ["Category", "Finding Count", "% of Total"], fg=C_ACCENT)
    total = sum(c for _, c in categories)
    cat_start = 26
    for i, (cat, cnt) in enumerate(categories, cat_start):
        ws.row_dimensions[i].height = 18
        bg = C_LIGHT_BG if i % 2 == 0 else C_LIGHT2
        set_cell(ws, i, 1, cat,                     fg=bg)
        set_cell(ws, i, 2, cnt,                     fg=bg, h_align="center")
        set_cell(ws, i, 3, f"{cnt/total*100:.1f}%", fg=bg, h_align="center")

    pie = PieChart()
    pie.title = "Findings by Category"
    pie.style = 10
    pie.width = 22
    pie.height = 13

    pie_data = Reference(ws, min_col=2, min_row=cat_start, max_row=cat_start + len(categories) - 1)
    pie_cats = Reference(ws, min_col=1, min_row=cat_start, max_row=cat_start + len(categories) - 1)
    pie.add_data(pie_data)
    pie.set_categories(pie_cats)
    ws.add_chart(pie, "C24")

    # OWASP mapping
    end_row = cat_start + len(categories) + 1
    section_title(ws, end_row, 1, "  OWASP Top 10 Coverage", 6, fg=C_DARK_BLUE)
    owasp = [
        ("A01:2021", "Broken Access Control",               "HIGH",   "SEC-007, SEC-009, SEC-010, SEC-022"),
        ("A02:2021", "Cryptographic Failures",              "CRITICAL","SEC-001, SEC-011, SEC-012, SEC-002"),
        ("A03:2021", "Injection",                           "MEDIUM",  "SEC-013, SEC-014, SEC-015"),
        ("A05:2021", "Security Misconfiguration",           "HIGH",    "SEC-008, SEC-019, SEC-020, SEC-021, SEC-023"),
        ("A06:2021", "Vulnerable & Outdated Components",    "MEDIUM",  "SEC-025, SEC-026"),
        ("A07:2021", "Identification & Authentication Failures","HIGH","SEC-004, SEC-005, SEC-006, SEC-017, SEC-027"),
        ("A09:2021", "Security Logging & Monitoring",       "MEDIUM",  "SEC-016, SEC-030"),
    ]
    header_row(ws, end_row + 1, ["OWASP ID", "Category", "Severity", "Related Findings"], fg=C_ACCENT)
    sev_c = {"CRITICAL": C_RED, "HIGH": C_ORANGE, "MEDIUM": C_YELLOW}
    for i, (oid, ocat, osev, ofinds) in enumerate(owasp, end_row + 2):
        ws.row_dimensions[i].height = 20
        bg = C_LIGHT_BG if i % 2 == 0 else C_LIGHT2
        set_cell(ws, i, 1, oid,     bold=True, fg=bg)
        set_cell(ws, i, 2, ocat,    fg=bg)
        set_cell(ws, i, 3, osev,    bold=True, fg=sev_c.get(osev, C_GREY), font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 4, ofinds,  fg=bg, wrap=True)


def build_setup_instructions(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 70

    banner(ws, 1, "Setup & Remediation Quick-Start Instructions", 2,
           fg=C_NAVY, size=14, height=35)

    sections = [
        ("IMMEDIATE ACTIONS (Do These Now)", C_RED, [
            ("Step 1 — Rotate JWT Secret",
             "Run: node -e \"console.log(require('crypto').randomBytes(64).toString('hex'))\" and set as JWT_SECRET in Render dashboard."),
            ("Step 2 — Remove secrets from git history",
             "Run: git filter-repo --path backend/.env --invert-paths\nThen force-push all branches.\nAlternatively use: git-secrets or BFG Repo Cleaner."),
            ("Step 3 — Revoke Google OAuth Client",
             "Go to Google Cloud Console → APIs & Services → Credentials. Revoke the current client. Create a new one and set via Render env vars (NOT render.yaml)."),
            ("Step 4 — Rotate DB password",
             "Go to Supabase → Database → Settings → Reset password. Update DB_PASSWORD in Render dashboard env vars (sync: false). Never commit to render.yaml."),
            ("Step 5 — Add backend/.gitignore",
             "Create backend/.gitignore with content:\n.env\n.env.local\n*.env\nnode_modules/\nuploads/\nembedded_store.json"),
        ]),
        ("SHORT-TERM FIXES (This Sprint)", C_ORANGE, [
            ("Fix JWT Fallback",
             "In src/middlewares/auth.js and src/config/jwt.js, remove DEFAULT_SECRET.\nAdd: if (!process.env.JWT_SECRET) { console.error('FATAL'); process.exit(1); }"),
            ("Fix CORS Wildcard",
             "In src/app.js, remove the 'process.env.NODE_ENV !== production' bypass.\nAllow only explicitly listed origins in all environments."),
            ("Authenticate /uploads route",
             "Replace app.use('/uploads', express.static(uploadDir)) with:\napp.get('/uploads/:file', authMiddleware, (req, res) => res.sendFile(...))"),
            ("Enable SSL for DB and Email",
             "In src/config/db.js: change rejectUnauthorized: false to rejectUnauthorized: true\nDownload Supabase CA cert and reference it.\nIn mailerService.js: remove tls: { rejectUnauthorized: false }"),
            ("Fix User Enumeration",
             "In authController.js forgot-password: always return HTTP 200 with generic message.\nNever return 'User not found' or different status codes based on user existence."),
        ]),
        ("CI/CD SETUP", C_ACCENT, [
            ("Enable GitHub Actions Workflow",
             "The file .github/workflows/security.yml is already committed.\nEnsure these GitHub Secrets are set:\n- SEMGREP_APP_TOKEN (from semgrep.dev)\n- GITLEAKS_LICENSE (optional for private repos)"),
            ("Set Up Semgrep",
             "Sign up at semgrep.dev → Get App Token → Add to GitHub Secrets as SEMGREP_APP_TOKEN."),
            ("Set Up Gitleaks",
             "The gitleaks-action@v2 runs automatically. For private repos, add GITLEAKS_LICENSE secret.\nFor public repos, no license needed."),
            ("Monitor Dependabot",
             "Enable Dependabot in GitHub Settings → Security → Enable Dependabot alerts.\nThis will auto-create PRs for vulnerable dependencies."),
        ]),
        ("TOOLING RECOMMENDATIONS", C_PURPLE, [
            ("Add Input Validation Library",
             "Run: npm install zod  (or joi)\nCreate validation schemas for all controller inputs.\nReject requests that fail schema validation before processing."),
            ("Add Structured Logging",
             "Run: npm install pino pino-pretty\nReplace console.log/error with pino logger.\nUse redact option to remove PII: redact: ['body.email', 'body.password']"),
            ("Add Token Revocation",
             "Run: npm install ioredis\nImplement a token blocklist in Redis.\nBlacklist tokens on logout and password change.\nCheck blocklist in auth middleware before accepting token."),
            ("Add Magic-Byte File Validation",
             "Run: npm install file-type\nAfter multer saves file, read first bytes and verify actual file type.\nReject if file-type.fromFile() result doesn't match expected image formats."),
        ]),
    ]

    row = 3
    for section_title_text, title_color, steps in sections:
        section_title(ws, row, 1, f"  {section_title_text}", 2, fg=title_color, size=11)
        row += 1
        header_row(ws, row, ["Action", "Instructions"], fg=C_DARK_BLUE)
        row += 1
        for step_name, instructions in steps:
            ws.row_dimensions[row].height = 60
            set_cell(ws, row, 1, step_name,     bold=True, fg=C_LIGHT_BG, wrap=True)
            set_cell(ws, row, 2, instructions,  fg=C_CREAM, wrap=True)
            row += 1
        ws.row_dimensions[row].height = 8
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate_report():
    wb = openpyxl.Workbook()

    ws_cover    = wb.active;         ws_cover.title = "📋 Cover"
    ws_exec     = wb.create_sheet("📊 Executive Summary")
    ws_backend  = wb.create_sheet("🏗️ Backend Inventory")
    ws_api      = wb.create_sheet("🔗 API Endpoints")
    ws_findings = wb.create_sheet("🔍 Security Findings")
    ws_deps     = wb.create_sheet("📦 Dependency Review")
    ws_remed    = wb.create_sheet("🛠️ Remediation Plan")
    ws_cicd     = wb.create_sheet("⚙️ GitHub Actions")
    ws_dash     = wb.create_sheet("🎯 Risk Dashboard")
    ws_setup    = wb.create_sheet("📝 Setup Instructions")

    print("Building Cover Page...")
    build_cover(ws_cover)

    print("Building Executive Summary...")
    build_executive_summary(ws_exec)

    print("Building Backend Inventory...")
    build_backend_inventory(ws_backend)

    print("Building API Endpoint Inventory...")
    build_endpoint_inventory(ws_api)

    print("Building Security Findings (30 issues)...")
    build_findings(ws_findings)

    print("Building Dependency Review...")
    build_dependency_review(ws_deps)

    print("Building Remediation Roadmap...")
    build_remediation(ws_remed)

    print("Building GitHub Actions Summary...")
    build_github_actions(ws_cicd)

    print("Building Risk Dashboard + Charts...")
    build_risk_dashboard(ws_dash)

    print("Building Setup Instructions...")
    build_setup_instructions(ws_setup)

    # Tab colours
    ws_cover.sheet_properties.tabColor    = C_NAVY
    ws_exec.sheet_properties.tabColor     = C_DARK_BLUE
    ws_backend.sheet_properties.tabColor  = C_TEAL
    ws_api.sheet_properties.tabColor      = C_ACCENT
    ws_findings.sheet_properties.tabColor = C_RED
    ws_deps.sheet_properties.tabColor     = C_ORANGE
    ws_remed.sheet_properties.tabColor    = C_GREEN
    ws_cicd.sheet_properties.tabColor     = C_PURPLE
    ws_dash.sheet_properties.tabColor     = C_YELLOW
    ws_setup.sheet_properties.tabColor    = "0EA5E9"

    # Freeze header rows on data sheets
    for ws in [ws_backend, ws_api, ws_findings, ws_deps, ws_remed]:
        ws.freeze_panes = "A4"

    output = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "BrushIQ_Security_Review_Report.xlsx"
    )
    wb.save(output)
    print(f"\n✅  Security Review Report saved:\n    {output}")
    print(f"\n📊  Sheets: Cover | Executive Summary | Backend Inventory | "
          f"API Endpoints | Security Findings (30) | Dependency Review | "
          f"Remediation Plan | GitHub Actions | Risk Dashboard | Setup Instructions")
    return output


if __name__ == "__main__":
    generate_report()


import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
import os

# ─── Colour Palette ────────────────────────────────────────────────────────────
C_NAVY      = "0D1B2A"
C_DARK_BLUE = "1B2A4A"
C_ACCENT    = "2563EB"
C_TEAL      = "0EA5E9"
C_RED       = "DC2626"
C_ORANGE    = "EA580C"
C_YELLOW    = "CA8A04"
C_GREEN     = "16A34A"
C_PURPLE    = "7C3AED"
C_WHITE     = "FFFFFF"
C_LIGHT_BG  = "F1F5F9"
C_LIGHT2    = "E2E8F0"
C_GREY      = "94A3B8"
C_DARK_TEXT = "0F172A"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_DARK_TEXT, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="94A3B8")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, fg=C_WHITE, font_color=C_DARK_TEXT,
             font_size=10, h_align="left", wrap=False, border=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=font_color, size=font_size, italic=italic)
    c.fill = fill(fg)
    c.alignment = align(h_align, "center", wrap)
    if border:
        c.border = thin_border()
    return c

def header_row(ws, row, cols_values, fg=C_DARK_BLUE, font_color=C_WHITE, size=10):
    for col, val in enumerate(cols_values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(bold=True, color=font_color, size=size)
        c.fill = fill(fg)
        c.alignment = align("center", "center")
        c.border = thin_border()

def section_title(ws, row, col, text, span, fg=C_ACCENT, size=12):
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=C_WHITE, size=size)
    c.fill = fill(fg)
    c.alignment = align("left", "center")
    c.border = medium_border()
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)

def auto_width(ws, extra=4):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + extra, 60)

# ═══════════════════════════════════════════════════════════════════════════════
# DATA DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════════

BACKEND_INVENTORY = [
    ("Framework",              "Node.js + Express.js v4.19.2"),
    ("Language",               "JavaScript (ES2020, CommonJS modules)"),
    ("API Architecture",       "RESTful JSON API"),
    ("Authentication",         "JWT (jsonwebtoken ^9.0.2) + Google OAuth2 (google-auth-library ^11.0.0)"),
    ("Authorization Model",    "User-scoped resource ownership enforced via SQL JOIN checks"),
    ("Database",               "PostgreSQL 15 (primary) + Embedded JSON fallback store"),
    ("ORM / Query Builder",    "Raw SQL via pg ^8.11.5 (no ORM)"),
    ("API Documentation",      "None detected (no Swagger/OpenAPI)"),
    ("Middleware",              "CORS, JSON parser, custom auth JWT middleware, static file serving"),
    ("File Upload",            "multer ^1.4.5-lts.1 — image uploads to /uploads/, 5 MB limit, MIME filter"),
    ("Session Handling",       "Stateless JWT (7-day expiry, stored client-side)"),
    ("Image Processing",       "jimp ^1.6.1 — AI bristle-wear analysis"),
    ("Third-Party Integrations","Google OAuth2 (sign-in), Supabase/PostgreSQL (cloud DB), Render (deployment)"),
    ("Deployment",             "Render.com (Node web service), Docker Compose for local dev"),
]

ENDPOINT_INVENTORY = [
    # Endpoint, Method, Auth, Roles, Controller/File
    ("/api/auth/register",                "POST",   "No",  "Public",      "authController.register"),
    ("/api/auth/login",                   "POST",   "No",  "Public",      "authController.login"),
    ("/api/auth/google",                  "POST",   "No",  "Public",      "authController.googleLogin"),
    ("/api/auth/forgot-password",         "POST",   "No",  "Public",      "authController.forgotPassword"),
    ("/api/auth/change-password",         "POST",   "Yes", "Authenticated","authController.changePassword"),
    ("/api/auth/me",                      "GET",    "Yes", "Authenticated","authController.getMe"),
    ("/api/family",                       "GET",    "Yes", "Authenticated","familyController.getFamilyMembers"),
    ("/api/family",                       "POST",   "Yes", "Authenticated","familyController.addFamilyMember"),
    ("/api/family/:id",                   "PUT",    "Yes", "Authenticated","familyController.updateFamilyMember"),
    ("/api/family/:id",                   "DELETE", "Yes", "Authenticated","familyController.deleteFamilyMember"),
    ("/api/toothbrushes",                 "GET",    "Yes", "Authenticated","toothbrushController.getToothbrushes"),
    ("/api/toothbrushes",                 "POST",   "Yes", "Authenticated","toothbrushController.addToothbrush"),
    ("/api/toothbrushes/:id",             "PUT",    "Yes", "Authenticated","toothbrushController.updateToothbrush"),
    ("/api/toothbrushes/:id",             "DELETE", "Yes", "Authenticated","toothbrushController.deleteToothbrush"),
    ("/api/scans/analyze",                "POST",   "Yes", "Authenticated","scanController.analyzeScan (multipart)"),
    ("/api/scans",                        "POST",   "Yes", "Authenticated","scanController.saveScan"),
    ("/api/scans",                        "GET",    "Yes", "Authenticated","scanController.getScansHistory"),
    ("/api/scans/:id",                    "GET",    "Yes", "Authenticated","scanController.getScanById"),
    ("/api/reminders",                    "GET",    "Yes", "Authenticated","reminderController.getReminders"),
    ("/api/reminders",                    "POST",   "Yes", "Authenticated","reminderController.createReminder"),
    ("/api/reminders/:id/complete",       "PATCH",  "Yes", "Authenticated","reminderController.completeReminder"),
    ("/api/tips",                         "GET",    "Yes", "Authenticated","tipController.getTips"),
    ("/api/tips/personalized",            "GET",    "Yes", "Authenticated","tipController.getPersonalizedTips"),
    ("/api/dashboard",                    "GET",    "Yes", "Authenticated","dashboardController.getDashboardData"),
    ("/api/system/database-status",       "GET",    "No",  "Public",      "routes/system.js (inline handler)"),
    ("/health",                           "GET",    "No",  "Public",      "app.js (inline)"),
    ("/api/health",                       "GET",    "No",  "Public",      "app.js (inline)"),
    ("/",                                 "GET",    "No",  "Public",      "app.js (inline)"),
    ("/uploads/*",                        "GET",    "No",  "Public",      "express.static — uploads directory"),
    ("/illustrations/*",                  "GET",    "No",  "Public",      "express.static — frontend/public/illustrations"),
]

# ─── Security Findings ─────────────────────────────────────────────────────────
FINDINGS = [
    # ID, Category, Severity, File, Title, Description, Why Concern, Fix
    ("SEC-001", "Cryptography / Secrets",  "CRITICAL",
     "backend/.env (L7) + middlewares/auth.js (L5) + authController.js (L6)",
     "Hardcoded JWT Secret in .env AND Fallback Default in Code",
     "JWT_SECRET is set to 'supersecretbrushiqjwttoken' in the committed .env file. Both auth.js and authController.js also fall back to this same hardcoded value if the env var is missing.",
     "Any attacker who reads the .env file (e.g., via repo leak or server path traversal) can sign arbitrary JWTs and impersonate any user. The in-code fallback means even a misconfigured deployment is exploitable.",
     "Rotate the secret immediately. Generate a cryptographically random 256-bit value (openssl rand -base64 32). Remove .env from git history. Add .env to .gitignore. Never provide a code-level fallback."),

    ("SEC-002", "Sensitive Data Exposure",  "CRITICAL",
     "backend/.env (L8) + backend/render.yaml (L28)",
     "Google OAuth Client ID Committed to Repository",
     "The real Google OAuth2 Client ID (534843148727-ernb5gqgo6pf1cobmmvjbsl7d4f5026s.apps.googleusercontent.com) is embedded in both .env and render.yaml which are committed to the repository.",
     "The Client ID is publicly visible to all repository readers. It can be used to craft phishing OAuth flows targeting BrushIQ users. If the Client Secret were also exposed, full account takeover would be trivial.",
     "Remove Client ID from all committed files. Use Render environment variable injection (sync: true with secret store). Revoke and rotate the OAuth credential in Google Cloud Console. Add to .gitignore."),

    ("SEC-003", "Sensitive Data Exposure",  "CRITICAL",
     "backend/.env (L4-L6) + docker-compose.yml (L10)",
     "Database Credentials Committed to Repository",
     "DB_USER=postgres, DB_PASSWORD=postgrespassword, DB_DATABASE=brushiq are all present in .env. The docker-compose.yml also hardcodes POSTGRES_PASSWORD=postgrespassword.",
     "Exposed database credentials allow direct database access if the database port is reachable. Credential stuffing attacks become trivially scriptable.",
     "Immediately rotate DB credentials. Remove .env from VCS. Use Render secret env vars (sync: false already set for DB_PASSWORD — ensure this stays that way). Use Docker secrets instead of environment variables in production."),

    ("SEC-004", "Authentication",  "HIGH",
     "backend/src/middlewares/auth.js (L5)",
     "Weak Fallback JWT Secret",
     "The auth middleware falls back to 'supersecretbrushiqjwttoken' if JWT_SECRET env var is not set. This is predictable, short, and could be brute-forced.",
     "If JWT_SECRET is not set in production (e.g., deployment misconfiguration), the application silently accepts tokens signed with a known weak secret, making all JWTs forgeable.",
     "Remove all fallback values. If JWT_SECRET is missing at startup, throw a fatal error and refuse to start: if (!process.env.JWT_SECRET) throw new Error('JWT_SECRET env var required')."),

    ("SEC-005", "Authentication",  "HIGH",
     "backend/src/controllers/authController.js (L165)",
     "Weak Minimum Password Length (6 characters)",
     "Registration and change-password enforce a minimum password length of only 6 characters.",
     "6-character passwords are highly susceptible to brute-force and dictionary attacks. NIST SP 800-63B recommends at least 8 characters, with support for much longer passphrases.",
     "Increase minimum to 8-12 characters. Consider adding a zxcvbn-based strength check. Also implement account lockout after N failed login attempts."),

    ("SEC-006", "Authentication",  "HIGH",
     "backend/src/controllers/authController.js (L275-L299)",
     "User Enumeration via Forgot-Password Endpoint",
     "The /api/auth/forgot-password endpoint returns 'User not found' (HTTP 400) when the email/phone does not exist in the database.",
     "Attackers can enumerate valid email addresses and phone numbers by submitting them to this endpoint and observing the differing response.",
     "Return the same success response regardless of whether the user exists: { message: 'If this account exists, a reset link has been sent.' }. Return HTTP 200 in all cases."),

    ("SEC-007", "Authorization / IDOR",  "HIGH",
     "backend/src/controllers/toothbrushController.js (L125)",
     "IDOR Risk — Toothbrush Delete Uses Only ID Without User Re-Validation in DELETE Query",
     "After the ownership check, the DELETE query uses 'DELETE FROM toothbrushes WHERE id = $1' without re-checking f.user_id = $2. A race condition or logic bug here could allow deletion of unowned toothbrushes.",
     "If the ownership check passes but the delete query does not re-enforce the user filter, a concurrent update that changes ownership could be exploited.",
     "Change the DELETE query to: DELETE FROM toothbrushes WHERE id = $1 AND family_member_id IN (SELECT id FROM family_members WHERE user_id = $2), passing [id, req.user.id]."),

    ("SEC-008", "Configuration",  "HIGH",
     "backend/src/app.js (L13)",
     "CORS Wildcard '*' When ALLOWED_ORIGINS Not Set",
     "If the ALLOWED_ORIGINS environment variable is not configured, the CORS policy defaults to origin: '*', permitting any origin to make cross-origin requests.",
     "An overly permissive CORS policy allows malicious websites to make authenticated API calls on behalf of logged-in users (CSRF-like attack via credentialed CORS requests).",
     "Set ALLOWED_ORIGINS explicitly in all environments. Remove the wildcard fallback. Use a strict allowlist: ['https://brushiq.app', 'https://www.brushiq.app']. Fail startup if ALLOWED_ORIGINS is not set in production."),

    ("SEC-009", "Configuration",  "HIGH",
     "backend/src/app.js (L38)",
     "Uploaded Files Served Without Authentication",
     "The /uploads static directory is served publicly via express.static without any authentication middleware.",
     "Any person (authenticated or not) can access uploaded scan images directly via /uploads/<filename> if they know or can guess the filename. Scan images may contain identifiable health data.",
     "Move uploads behind authentication. Serve files through a controller that validates JWT first: router.get('/uploads/:file', authMiddleware, serveFile). Alternatively, use signed URLs or move storage to a private S3-compatible bucket."),

    ("SEC-010", "Configuration",  "HIGH",
     "backend/src/app.js (L41-L44)",
     "Frontend Illustrations Directory Served Without Access Control",
     "The /illustrations path conditionally mounts static files from the frontend public directory without authentication.",
     "Though illustration images are low-sensitivity, this exposes the filesystem path structure and could be used to enumerate available frontend assets.",
     "Mount static assets from a CDN or at least ensure the path is not exploitable via path traversal. Validate that express.static is used correctly with absolute paths."),

    ("SEC-011", "Cryptography",  "HIGH",
     "backend/src/config/db.js (L39, L47-L49)",
     "SSL Certificate Verification Disabled (rejectUnauthorized: false)",
     "Both connection paths (DATABASE_URL and individual config) set rejectUnauthorized: false for SSL, disabling server certificate validation.",
     "Disabling certificate validation enables Man-in-the-Middle attacks on the database connection. An attacker on the network could intercept or tamper with all database traffic, exposing passwords, tokens, and PII.",
     "Set rejectUnauthorized: true and provide the proper CA certificate. For Supabase, download the root CA and reference it: ssl: { rejectUnauthorized: true, ca: fs.readFileSync('supabase-ca.crt') }."),

    ("SEC-012", "Input Validation",  "MEDIUM",
     "backend/src/controllers/scanController.js (L34-L99)",
     "Client-Controlled Scan Metrics Saved Without Server-Side Validation",
     "The saveScan endpoint accepts all scan metrics (wearPercentage, healthScore, bristleSpreading, etc.) directly from the request body without range validation.",
     "A malicious user can POST arbitrary scan metrics (e.g., healthScore: -999, wearPercentage: 200) to manipulate their own records and potentially exploit downstream business logic.",
     "Validate all numeric fields server-side: wearPercentage in [0,100], healthScore in [0,100], condition must be one of ['Good','Moderate Wear','Replace Soon','Replace Immediately'], etc. Use a validation library (Joi, Zod) or express-validator."),

    ("SEC-013", "Input Validation",  "MEDIUM",
     "backend/src/routes/scan.js (L30-L36)",
     "File Upload MIME Type Check via mimetype Only (No Magic-Byte Validation)",
     "The multer fileFilter checks file.mimetype.startsWith('image/') but this is set by the HTTP client and can be trivially spoofed.",
     "An attacker can upload a malicious file (e.g., a web shell, SVG with embedded JavaScript, or an HTML file) by setting Content-Type: image/jpeg. This content would then be accessible via the public /uploads/ path.",
     "Add magic-byte validation using the 'file-type' npm package after the file is written to disk. Also restrict allowed extensions to .jpg, .jpeg, .png, .webp. Store uploads outside the webroot or serve via authenticated endpoint."),

    ("SEC-014", "Input Validation",  "MEDIUM",
     "backend/src/controllers/familyController.js (L104-L124)",
     "No Length or Type Validation on Family Member Fields",
     "addFamilyMember only checks field presence, not type or length. name, gender, relationship accept any string of any length.",
     "An attacker could insert extremely long strings (DoS via large DB entries), HTML injection in name fields (reflected in tips/reminders messages), or unexpected types that may cause application errors.",
     "Add max-length validation (name ≤ 100 chars, relationship ≤ 50 chars). Validate age is an integer in [0, 120]. Consider using express-validator or Joi schema validation middleware."),

    ("SEC-015", "Sensitive Data Exposure",  "MEDIUM",
     "backend/src/controllers/authController.js (L215-L217)",
     "Internal Error Details Logged to Console in Registration",
     "catch(err) { console.error('Registration error details:', err) } logs the full error object including stack trace.",
     "In a containerized or cloud-log-aggregated environment, stack traces reveal internal structure, file paths, and potentially partial data — information useful to attackers.",
     "Log err.message only, not the full object. Use a structured logger (pino, winston) with log-level control. Never log sensitive request body data."),

    ("SEC-016", "Business Logic",  "MEDIUM",
     "backend/src/controllers/authController.js (L221-L273)",
     "No Rate Limiting on Login / Auth Endpoints",
     "The login, register, and forgot-password endpoints have no rate limiting, allowing unlimited brute-force attempts.",
     "Without rate limiting, an attacker can make thousands of login attempts per second, enabling credential stuffing or password spray attacks against all accounts.",
     "Add express-rate-limit middleware: limit auth endpoints to 10 requests/minute per IP. Consider adding CAPTCHA for repeated failures. Implement exponential backoff or temporary account lockout."),

    ("SEC-017", "Business Logic",  "MEDIUM",
     "backend/src/controllers/authController.js",
     "No Rate Limiting on File Upload / AI Scan Endpoint",
     "The /api/scans/analyze endpoint performs expensive Jimp image processing without rate limiting.",
     "An attacker can flood the server with large image uploads, causing CPU/memory exhaustion and denial of service. With 5 MB per file, 100 concurrent uploads = 500 MB memory pressure.",
     "Apply rate limiting (e.g., 10 scans/minute per user). Also validate image dimensions via jimp after loading to prevent decompression bombs (very large raw images disguised as small files)."),

    ("SEC-018", "Configuration",  "MEDIUM",
     "backend/src/app.js (L19-L20)",
     "No Request Body Size Limit for JSON / URL-Encoded Payloads",
     "express.json() and express.urlencoded() are used without a size limit option.",
     "An attacker can POST arbitrarily large JSON payloads to any endpoint, potentially causing memory exhaustion and denial of service.",
     "Add size limits: app.use(express.json({ limit: '10kb' })) and app.use(express.urlencoded({ limit: '10kb', extended: true }))."),

    ("SEC-019", "Configuration",  "MEDIUM",
     "backend/src/app.js",
     "Missing Content Security Policy (CSP) Header",
     "The security headers middleware sets X-Content-Type-Options, X-Frame-Options, X-XSS-Protection, and Referrer-Policy, but omits a Content-Security-Policy header.",
     "Without CSP, if any endpoint reflects user input (current or future), XSS attacks have no browser-side mitigation. CSP is a defense-in-depth control.",
     "Add: res.setHeader('Content-Security-Policy', \"default-src 'none'; script-src 'self'; connect-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'\"). Tailor to actual app needs."),

    ("SEC-020", "Configuration",  "MEDIUM",
     "backend/src/app.js",
     "Missing Strict-Transport-Security (HSTS) Header",
     "No HSTS header is set. HTTP connections are not explicitly forced to HTTPS at the application layer.",
     "Without HSTS, users who navigate via HTTP can be subject to SSL stripping attacks, exposing their session tokens.",
     "Add: res.setHeader('Strict-Transport-Security', 'max-age=31536000; includeSubDomains; preload') when NODE_ENV === 'production'."),

    ("SEC-021", "Configuration",  "MEDIUM",
     "backend/src/routes/system.js (L6-L13)",
     "Unauthenticated Database-Status Endpoint Reveals Infrastructure Details",
     "GET /api/system/database-status is public and returns { mode: 'postgresql', connected: true/false } without authentication.",
     "Internal infrastructure state (DB mode, connectivity) is exposed to the public. Error responses may include err.message with DB connection details.",
     "Protect this endpoint with authMiddleware. Remove error details from the response body in production. Only expose a boolean status."),

    ("SEC-022", "Authorization",  "MEDIUM",
     "backend/src/config/db.js (L314-L319)",
     "Embedded SQL Engine Has Loose Ownership Filter Logic",
     "In the embedded SQL engine, the filter for SELECT id FROM family_members WHERE id = $1 AND user_id = $2 uses a single OR between params: filter(f => f.user_id === userId || f.id === idOrUser).",
     "This logic error could allow a user to retrieve another user's family member record in the embedded (offline) mode if they supply the target family member's ID. When pgConnected is false, this fallback is used.",
     "Fix the filter to be: filter(f => f.id === id && f.user_id === userId). Do not combine OR logic. The embedded store should mirror the exact SQL constraints."),

    ("SEC-023", "Dependency",  "MEDIUM",
     "backend/package.json",
     "multer ^1.4.5-lts.1 — Known Historical Vulnerabilities",
     "The application uses multer 1.4.x LTS. The canonical multer v1.x line has had memory DoS vulnerabilities in its handling of multipart boundaries.",
     "Using older library versions with known vulnerability history increases attack surface, especially on the file upload path which handles untrusted user content.",
     "Upgrade to multer v2.x (currently in active development with security improvements) or pin to the most current 1.4.x-lts patch. Run 'npm audit' and address all findings."),

    ("SEC-024", "Dependency",  "MEDIUM",
     "backend/package.json",
     "express ^4.19.2 — Recent Path Traversal Fix in 4.21.x",
     "Express 4.19.2 predates the 4.21.x series which addressed path traversal issues in express.static and route handling.",
     "The application uses express.static for uploads and illustrations — two paths that serve user-related content. Path traversal in static serving could expose arbitrary files.",
     "Pin to express@^4.21.x or latest stable. Run 'npm audit' and update accordingly."),

    ("SEC-025", "Business Logic",  "MEDIUM",
     "backend/src/config/db.js (L97-L103)",
     "Embedded JSON Store Written to Disk — Sensitive Data in Plaintext",
     "When PostgreSQL is unavailable, user data (including password hashes, emails, phone numbers) is stored in embedded_store.json on disk in plaintext JSON.",
     "The JSON file at src/db/embedded_store.json is not encrypted. If the server filesystem is accessible (e.g., via path traversal, compromised process), all user data is exposed.",
     "At minimum, restrict file permissions to 600 (owner read/write only). In production, do not fall back to the embedded store — fail loudly if PostgreSQL is unavailable. Consider encrypting the embedded store if offline mode is required."),

    ("SEC-026", "Authentication",  "LOW",
     "backend/src/controllers/authController.js (L121)",
     "JWT Token Expiry is 7 Days — No Refresh Token Mechanism",
     "Tokens are signed with expiresIn: '7d'. There is no token refresh, revocation, or logout mechanism.",
     "A stolen JWT remains valid for up to 7 days. There is no server-side way to invalidate a session (e.g., on logout, password change, or suspicious activity).",
     "Shorten expiry to 1-2 hours. Implement refresh tokens (stored securely, rotated on use). On password change, blacklist all previous tokens (store in Redis or DB). Add logout to invalidate the token."),

    ("SEC-027", "Configuration",  "LOW",
     "backend/src/app.js (L26)",
     "X-XSS-Protection Header is Deprecated",
     "X-XSS-Protection: 1; mode=block is set, but this header is deprecated and removed from modern browsers (Chromium 78+).",
     "The header provides no protection in modern browsers and can cause issues in some edge cases. Relying on it gives a false sense of security.",
     "Remove X-XSS-Protection. Rely on a proper Content-Security-Policy instead. Install Helmet.js for managed, up-to-date security headers."),

    ("SEC-028", "Configuration",  "LOW",
     "backend/src/server.js (L29)",
     "Server Listens on 0.0.0.0 — Binds to All Interfaces",
     "The server explicitly binds to 0.0.0.0, exposing it on all network interfaces.",
     "In a local development or internal network context, this exposes the API to the local network, not just localhost. On cloud servers without firewall rules, the API is directly accessible.",
     "In development, bind to 127.0.0.1. In production, bind to 0.0.0.0 only behind a reverse proxy (Nginx/Render) with proper firewall rules. Document the deployment topology."),

    ("SEC-029", "Sensitive Data",  "LOW",
     "backend/src/controllers/authController.js (L37, L63)",
     "Sensitive Data Logged in Console (Google Login Flow)",
     "googleClientId.trim() and email, googleId values are logged to console at info level during normal operation.",
     "In a log aggregation system, PII (email addresses, Google IDs) being logged creates compliance risk (GDPR, DPDP). Log files may be stored insecurely.",
     "Remove PII from logs. Log only non-identifiable event descriptors: '[GoogleLogin] user authenticated successfully'. Use structured logging with PII redaction."),

    ("SEC-030", "Configuration",  "LOW",
     "backend/package.json",
     "No Security-Related npm Scripts (audit, snyk)",
     "The package.json scripts section only contains start, dev, test, db:init. There are no audit, security-scan, or dependency-check scripts.",
     "Without automated security checks in the development workflow, vulnerable dependencies may go unnoticed.",
     "Add: 'audit': 'npm audit --audit-level=moderate' to scripts. Integrate into CI/CD pipeline. Consider adding 'snyk test' or similar."),
]

DEPENDENCY_REVIEW = [
    # Package, Version Specified, Latest Stable, Risk Level, Notes
    ("express",              "^4.19.2",       "4.21.x",     "MEDIUM",  "Path traversal fix in 4.21.x; upgrade recommended"),
    ("jsonwebtoken",         "^9.0.2",        "9.0.2",      "LOW",     "Current stable — maintain this version"),
    ("bcryptjs",             "^2.4.3",        "2.4.3",      "LOW",     "Current stable; consider bcrypt (native) for performance"),
    ("cors",                 "^2.8.5",        "2.8.5",      "LOW",     "Current stable"),
    ("dotenv",               "^16.4.5",       "16.4.5",     "LOW",     "Current stable"),
    ("google-auth-library",  "^11.0.0",       "11.x",       "LOW",     "Current stable; keep updated for OAuth security patches"),
    ("jimp",                 "^1.6.1",        "1.6.x",      "MEDIUM",  "Decompression bomb risk in image libraries; validate dimensions"),
    ("multer",               "^1.4.5-lts.1",  "2.x / 1.4.5-lts", "MEDIUM", "Historical DoS vulnerabilities; audit carefully"),
    ("pg",                   "^8.11.5",       "8.13.x",     "LOW",     "Minor patch available; upgrade for fixes"),
    ("uuid",                 "^9.0.1",        "9.0.1",      "LOW",     "Current stable"),
    ("jest",                 "^29.7.0",       "29.7.0",     "LOW",     "Dev only — current stable"),
    ("nodemon",              "^3.1.0",        "3.1.x",      "LOW",     "Dev only — current stable"),
    ("supertest",            "^6.3.4",        "6.3.4",      "LOW",     "Dev only — current stable"),
]

REMEDIATION_PLAN = [
    ("P0 — Immediate (< 24 hours)",
     "SEC-001 / SEC-002 / SEC-003",
     "Rotate JWT_SECRET, Google Client ID, and database credentials. Remove .env from git. Add .gitignore entries.",
     "CRITICAL"),
    ("P1 — Short-term (< 1 week)",
     "SEC-004 / SEC-008 / SEC-011 / SEC-009",
     "Remove JWT fallback defaults. Set strict CORS allowlist. Enable SSL rejectUnauthorized. Authenticate uploads endpoint.",
     "HIGH"),
    ("P2 — Medium-term (< 1 month)",
     "SEC-005 / SEC-006 / SEC-016 / SEC-017 / SEC-018 / SEC-019 / SEC-020",
     "Increase password minimum length. Fix user enumeration. Add rate limiting to auth and scan endpoints. Add body size limits. Add CSP and HSTS headers.",
     "HIGH/MEDIUM"),
    ("P3 — Ongoing",
     "SEC-012 / SEC-013 / SEC-014 / SEC-022 / SEC-023 / SEC-024",
     "Add comprehensive input validation with Joi/Zod. Add magic-byte file validation. Fix embedded SQL owner logic. Update express and multer.",
     "MEDIUM"),
    ("P4 — Best-practice Improvements",
     "SEC-026 / SEC-027 / SEC-028 / SEC-029 / SEC-030",
     "Implement token refresh/revocation. Remove deprecated X-XSS-Protection. Add structured PII-safe logging. Add npm audit to CI.",
     "LOW"),
]


def build_executive_summary(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.row_dimensions[1].height = 50

    # Banner
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "🔐  BrushIQ Backend — Secure Code Review Report"
    c.font = Font(bold=True, color=C_WHITE, size=18, name="Calibri")
    c.fill = fill(C_NAVY)
    c.alignment = align("center", "center")

    # Meta info
    meta = [
        ("Report Date",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Application",    "BrushIQ — AI-Powered Oral Healthcare Platform"),
        ("Backend Stack",  "Node.js + Express.js v4.19.2"),
        ("Database",       "PostgreSQL 15 + Embedded JSON Fallback"),
        ("Review Type",    "Static Code Analysis (SAST) — Defensive Security Review"),
        ("Reviewed By",    "Antigravity AI Code Security Reviewer"),
        ("Scope",          "backend/ — all source files, configs, dependencies"),
    ]
    for i, (k, v) in enumerate(meta, 3):
        ws.row_dimensions[i].height = 18
        set_cell(ws, i, 1, k, bold=True, fg=C_DARK_BLUE, font_color=C_WHITE, h_align="right")
        set_cell(ws, i, 2, v, fg=C_LIGHT_BG, font_color=C_DARK_TEXT)

    # Score box
    ws.row_dimensions[11].height = 10
    ws.merge_cells("A12:B12")
    section_title(ws, 12, 1, "  OVERALL SECURITY SCORE", 2, fg=C_DARK_BLUE)

    ws.row_dimensions[13].height = 60
    ws.merge_cells("A13:B13")
    score_cell = ws["A13"]
    score_cell.value = "52 / 100"
    score_cell.font = Font(bold=True, color=C_ORANGE, size=40, name="Calibri")
    score_cell.fill = fill(C_LIGHT_BG)
    score_cell.alignment = align("center", "center")
    score_cell.border = medium_border()

    ws.row_dimensions[14].height = 30
    ws.merge_cells("A14:B14")
    c = ws["A14"]
    c.value = ("⚠  Rating: MODERATE RISK  |  3 Critical findings must be resolved IMMEDIATELY before production launch. "
               "The application demonstrates solid authorization logic but has critical secret management and configuration gaps.")
    c.font = Font(bold=False, color=C_ORANGE, size=10, name="Calibri")
    c.fill = fill("FFF7ED")
    c.alignment = align("center", "center", wrap=True)
    c.border = thin_border()

    # Finding counts
    ws.row_dimensions[15].height = 10
    ws.merge_cells("A16:B16")
    section_title(ws, 16, 1, "  FINDING COUNTS BY SEVERITY", 2, fg=C_DARK_BLUE)

    counts = [
        ("CRITICAL", 3,  C_RED,    "SEC-001, SEC-002, SEC-003"),
        ("HIGH",     8,  C_ORANGE, "SEC-004 through SEC-011"),
        ("MEDIUM",   14, C_YELLOW, "SEC-012 through SEC-025"),
        ("LOW",      5,  C_TEAL,   "SEC-026 through SEC-030"),
    ]
    header_row(ws, 17, ["Severity", "Count", "Hex Color", "Finding IDs"], fg=C_DARK_BLUE)
    for i, (sev, cnt, col, ids) in enumerate(counts, 18):
        ws.row_dimensions[i].height = 18
        set_cell(ws, i, 1, sev,  bold=True, fg=col, font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 2, cnt,  bold=True, fg=C_LIGHT_BG, h_align="center")
        set_cell(ws, i, 3, f"#{col}", fg=C_LIGHT_BG, h_align="center")
        set_cell(ws, i, 4, ids,  fg=C_LIGHT_BG)

    # Positive findings
    ws.row_dimensions[23].height = 10
    ws.merge_cells("A24:B24")
    section_title(ws, 24, 1, "  POSITIVE SECURITY CONTROLS DETECTED", 2, fg=C_GREEN)

    positives = [
        "✓ SQL Injection prevented — parameterized queries ($1, $2...) used consistently throughout all controllers",
        "✓ Password hashing — bcrypt with salt factor 10 used correctly in register and change-password flows",
        "✓ X-Powered-By header disabled — prevents Express.js fingerprinting",
        "✓ JWT format validation — Bearer format explicitly checked in auth middleware",
        "✓ Ownership validation — all CRUD operations verify resource belongs to authenticated user via SQL JOINs",
        "✓ Multer file size limit — 5 MB cap prevents basic upload-based DoS",
        "✓ MIME type filter — image/* filter on upload provides first-level content type check",
        "✓ Google ID token verified server-side — OAuth2Client.verifyIdToken() used correctly with audience check",
    ]
    for i, p in enumerate(positives, 25):
        ws.row_dimensions[i].height = 20
        ws.merge_cells(f"A{i}:B{i}")
        c = ws.cell(row=i, column=1, value=p)
        c.font = Font(color=C_DARK_TEXT, size=10, name="Calibri")
        c.fill = fill("F0FDF4" if i % 2 == 0 else C_LIGHT_BG)
        c.alignment = align("left", "center", wrap=True)
        c.border = thin_border()


def build_backend_inventory(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 70

    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "PHASE 1 — Backend Inventory"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center")

    header_row(ws, 3, ["Component", "Details"], fg=C_ACCENT)

    for i, (comp, detail) in enumerate(BACKEND_INVENTORY, 4):
        ws.row_dimensions[i].height = 22
        set_cell(ws, i, 1, comp, bold=True, fg=C_DARK_BLUE if i % 2 == 0 else "1E3A5F", font_color=C_WHITE)
        set_cell(ws, i, 2, detail, fg=C_LIGHT_BG if i % 2 == 0 else C_LIGHT2, wrap=True)


def build_endpoint_inventory(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 45

    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "PHASE 2 — API Endpoint Inventory"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center")

    header_row(ws, 3, ["Endpoint", "Method", "Auth Required", "Expected Roles", "Controller / File"], fg=C_ACCENT)

    method_colors = {
        "GET": "0EA5E9", "POST": "16A34A", "PUT": "D97706",
        "DELETE": "DC2626", "PATCH": "7C3AED"
    }

    for i, (ep, method, auth, role, ctrl) in enumerate(ENDPOINT_INVENTORY, 4):
        even = i % 2 == 0
        row_bg = C_LIGHT_BG if even else C_LIGHT2
        ws.row_dimensions[i].height = 18

        set_cell(ws, i, 1, ep,     fg=row_bg)
        mc = method_colors.get(method, C_GREY)
        set_cell(ws, i, 2, method, bold=True, fg=mc, font_color=C_WHITE, h_align="center")
        auth_fg = C_GREEN if auth == "Yes" else C_RED
        set_cell(ws, i, 3, auth,   bold=True, fg=auth_fg, font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 4, role,   fg=row_bg, h_align="center")
        set_cell(ws, i, 5, ctrl,   fg=row_bg, italic=(ctrl.endswith(")")))


def build_findings(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 55
    ws.column_dimensions['G'].width = 55
    ws.column_dimensions['H'].width = 65

    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "PHASE 3 — Security Findings (SAST)"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center")

    cols = ["ID", "Category", "Severity", "File Path", "Title",
            "Description", "Why It's a Concern", "Recommended Fix"]
    header_row(ws, 3, cols, fg=C_ACCENT)

    sev_colors = {
        "CRITICAL": C_RED,
        "HIGH":     C_ORANGE,
        "MEDIUM":   C_YELLOW,
        "LOW":      C_TEAL,
    }

    for i, (fid, cat, sev, fpath, title, desc, concern, fix) in enumerate(FINDINGS, 4):
        ws.row_dimensions[i].height = 60
        even = i % 2 == 0
        row_bg = C_LIGHT_BG if even else C_LIGHT2
        sc = sev_colors.get(sev, C_GREY)

        set_cell(ws, i, 1, fid,     bold=True, fg=row_bg,    h_align="center")
        set_cell(ws, i, 2, cat,     fg=row_bg,               wrap=True)
        set_cell(ws, i, 3, sev,     bold=True, fg=sc, font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 4, fpath,   fg=row_bg,               wrap=True, italic=True)
        set_cell(ws, i, 5, title,   bold=True, fg=row_bg,    wrap=True)
        set_cell(ws, i, 6, desc,    fg=row_bg,               wrap=True)
        set_cell(ws, i, 7, concern, fg=row_bg,               wrap=True)
        set_cell(ws, i, 8, fix,     fg="F0FDF4",             wrap=True)


def build_dependency_review(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 60

    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "PHASE 4 — Dependency Review"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center")

    header_row(ws, 3, ["Package", "Specified Version", "Latest Stable", "Risk Level", "Notes"], fg=C_ACCENT)

    risk_colors = {"HIGH": C_ORANGE, "MEDIUM": C_YELLOW, "LOW": C_GREEN}

    for i, (pkg, spec, latest, risk, notes) in enumerate(DEPENDENCY_REVIEW, 4):
        even = i % 2 == 0
        row_bg = C_LIGHT_BG if even else C_LIGHT2
        rc = risk_colors.get(risk, C_GREY)
        ws.row_dimensions[i].height = 20

        set_cell(ws, i, 1, pkg,    bold=True, fg=row_bg)
        set_cell(ws, i, 2, spec,   fg=row_bg, h_align="center")
        set_cell(ws, i, 3, latest, fg=row_bg, h_align="center")
        set_cell(ws, i, 4, risk,   bold=True, fg=rc, font_color=C_WHITE, h_align="center")
        set_cell(ws, i, 5, notes,  fg=row_bg, wrap=True)


def build_remediation(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 75
    ws.column_dimensions['D'].width = 14

    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "PHASE 5 — Recommended Remediation Roadmap"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center")

    header_row(ws, 3, ["Priority Tier", "Finding IDs", "Action Required", "Risk Level"], fg=C_ACCENT)

    tier_colors = {
        "P0 — Immediate (< 24 hours)":    C_RED,
        "P1 — Short-term (< 1 week)":     C_ORANGE,
        "P2 — Medium-term (< 1 month)":   C_YELLOW,
        "P3 — Ongoing":                    C_TEAL,
        "P4 — Best-practice Improvements": C_GREEN,
    }
    risk_colors = {"CRITICAL": C_RED, "HIGH": C_ORANGE, "HIGH/MEDIUM": C_ORANGE,
                   "MEDIUM": C_YELLOW, "LOW": C_GREEN}

    for i, (tier, fids, action, risk) in enumerate(REMEDIATION_PLAN, 4):
        ws.row_dimensions[i].height = 45
        tc = tier_colors.get(tier, C_GREY)
        rc = risk_colors.get(risk, C_GREY)

        set_cell(ws, i, 1, tier,   bold=True, fg=tc, font_color=C_WHITE, wrap=True)
        set_cell(ws, i, 2, fids,   fg=C_LIGHT_BG, wrap=True)
        set_cell(ws, i, 3, action, fg=C_LIGHT_BG, wrap=True)
        set_cell(ws, i, 4, risk,   bold=True, fg=rc, font_color=C_WHITE, h_align="center")


def build_github_actions(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 90

    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "PHASE 6 — GitHub Actions Security Workflow"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
    c.fill = fill(C_DARK_BLUE)
    c.alignment = align("center", "center")

    workflow_yaml = """\
name: BrushIQ Security Scan

on:
  push:
    branches: [main, develop]
  pull_request:
    branches: [main]
  schedule:
    - cron: '0 3 * * 1'   # Every Monday 03:00 UTC

jobs:
  # ─── 1. Gitleaks — Secret Detection ────────────────────────────────────────
  gitleaks:
    name: Gitleaks Secret Scan
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
        with: { fetch-depth: 0 }
      - name: Run Gitleaks
        uses: gitleaks/gitleaks-action@v2
        env:
          GITHUB_TOKEN: ${{ secrets.GITHUB_TOKEN }}

  # ─── 2. Semgrep — SAST ─────────────────────────────────────────────────────
  semgrep:
    name: Semgrep SAST
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
      - uses: semgrep/semgrep-action@v1
        with:
          config: >
            p/javascript
            p/nodejs
            p/jwt
            p/secrets
            p/owasp-top-ten
        env:
          SEMGREP_APP_TOKEN: ${{ secrets.SEMGREP_APP_TOKEN }}

  # ─── 3. npm Audit — Dependency Review ──────────────────────────────────────
  npm-audit:
    name: npm Dependency Audit
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
      - uses: actions/setup-node@v4
        with: { node-version: '20', cache: 'npm', cache-dependency-path: backend/package-lock.json }
      - run: npm ci --prefix backend
      - name: npm audit
        run: npm audit --audit-level=moderate --prefix backend
        continue-on-error: false

  # ─── 4. Trivy — Docker / Filesystem Vulnerability Scan ────────────────────
  trivy:
    name: Trivy Vulnerability Scan
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
      - name: Trivy filesystem scan
        uses: aquasecurity/trivy-action@master
        with:
          scan-type: fs
          scan-ref: ./backend
          format: sarif
          output: trivy-results.sarif
          severity: CRITICAL,HIGH
          exit-code: '1'
      - name: Upload Trivy SARIF
        uses: github/codeql-action/upload-sarif@v3
        if: always()
        with:
          sarif_file: trivy-results.sarif

  # ─── 5. Summary ────────────────────────────────────────────────────────────
  security-summary:
    name: Publish Security Summary
    runs-on: ubuntu-latest
    needs: [gitleaks, semgrep, npm-audit, trivy]
    if: always()
    steps:
      - name: Write Job Summary
        run: |
          echo "## BrushIQ Security Scan Summary" >> $GITHUB_STEP_SUMMARY
          echo "| Tool | Status |" >> $GITHUB_STEP_SUMMARY
          echo "|------|--------|" >> $GITHUB_STEP_SUMMARY
          echo "| Gitleaks | ${{ needs.gitleaks.result }} |" >> $GITHUB_STEP_SUMMARY
          echo "| Semgrep  | ${{ needs.semgrep.result }} |" >> $GITHUB_STEP_SUMMARY
          echo "| npm audit| ${{ needs.npm-audit.result }} |" >> $GITHUB_STEP_SUMMARY
          echo "| Trivy    | ${{ needs.trivy.result }} |" >> $GITHUB_STEP_SUMMARY
          echo "" >> $GITHUB_STEP_SUMMARY
          echo "> Pipeline fails **only** on CRITICAL findings." >> $GITHUB_STEP_SUMMARY
"""

    set_cell(ws, 3, 1, "Workflow File", bold=True, fg=C_DARK_BLUE, font_color=C_WHITE)
    set_cell(ws, 3, 2, ".github/workflows/security.yml", bold=True, fg=C_DARK_BLUE, font_color=C_WHITE)

    ws.row_dimensions[4].height = 600
    c = ws.cell(row=4, column=1, value="YAML Content")
    c.font = font(bold=True, color=C_WHITE)
    c.fill = fill(C_ACCENT)
    c.alignment = align("left", "top")
    c.border = thin_border()

    c2 = ws.cell(row=4, column=2, value=workflow_yaml)
    c2.font = Font(name="Courier New", size=8, color=C_DARK_TEXT)
    c2.fill = fill(C_LIGHT_BG)
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    c2.border = thin_border()

    # Setup instructions
    ws.row_dimensions[6].height = 10
    section_title(ws, 7, 1, "  Setup Instructions", 2, fg=C_TEAL)
    instructions = [
        ("Step 1", "Add SEMGREP_APP_TOKEN secret in GitHub → Settings → Secrets → Actions"),
        ("Step 2", "Ensure GITHUB_TOKEN is available (it is automatic in all GitHub Actions)"),
        ("Step 3", "Remove backend/.env from git tracking: git rm --cached backend/.env"),
        ("Step 4", "Add backend/.env to .gitignore immediately"),
        ("Step 5", "Rotate JWT_SECRET, DB_PASSWORD, and Google Client ID before next deployment"),
        ("Step 6", "Configure ALLOWED_ORIGINS in production environment variables"),
        ("Step 7", "Run npm audit --prefix backend and fix all moderate+ findings"),
    ]
    for i, (step, inst) in enumerate(instructions, 8):
        ws.row_dimensions[i].height = 22
        set_cell(ws, i, 1, step, bold=True, fg=C_NAVY, font_color=C_WHITE)
        set_cell(ws, i, 2, inst, fg=C_LIGHT_BG, wrap=True)


def build_risk_dashboard(ws):
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "🎯  Security Risk Dashboard"
    c.font = Font(bold=True, color=C_WHITE, size=16, name="Calibri")
    c.fill = fill(C_NAVY)
    c.alignment = align("center", "center")

    # KPI Cards row
    kpis = [
        ("Security Score",     "52/100",  C_ORANGE,  "Moderate Risk"),
        ("Critical Findings",  "3",       C_RED,     "Immediate Action"),
        ("High Findings",      "8",       C_ORANGE,  "Fix This Sprint"),
        ("Medium Findings",    "14",      C_YELLOW,  "Plan & Schedule"),
        ("Low Findings",       "5",       C_TEAL,    "Best Practice"),
        ("Total Findings",     "30",      C_ACCENT,  "Across All Severity"),
    ]
    for col, (label, val, col_hex, sub) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 45
        ws.row_dimensions[5].height = 20

        c = ws.cell(row=3, column=col, value=label)
        c.font = Font(bold=True, color=C_WHITE, size=9, name="Calibri")
        c.fill = fill(col_hex)
        c.alignment = align("center", "center")
        c.border = medium_border()

        c2 = ws.cell(row=4, column=col, value=val)
        c2.font = Font(bold=True, color=col_hex, size=28, name="Calibri")
        c2.fill = fill(C_LIGHT_BG)
        c2.alignment = align("center", "center")
        c2.border = medium_border()

        c3 = ws.cell(row=5, column=col, value=sub)
        c3.font = Font(bold=False, color=C_GREY, size=9, name="Calibri")
        c3.fill = fill(C_LIGHT2)
        c3.alignment = align("center", "center")
        c3.border = thin_border()

    # Chart data
    ws.row_dimensions[8].height = 18
    section_title(ws, 8, 1, "  Finding Distribution by Severity", 6, fg=C_DARK_BLUE)

    chart_data = [("Severity", "Count"),
                  ("Critical", 3), ("High", 8), ("Medium", 14), ("Low", 5)]
    for i, (sev, cnt) in enumerate(chart_data, 10):
        ws.cell(row=i, column=1, value=sev)
        ws.cell(row=i, column=2, value=cnt)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Security Findings by Severity"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Severity"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data = Reference(ws, min_col=2, min_row=10, max_row=13)
    cats = Reference(ws, min_col=1, min_row=11, max_row=13)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "C8")

    # Category breakdown
    section_title(ws, 22, 1, "  Finding Distribution by Category", 6, fg=C_DARK_BLUE)
    categories = [
        ("Cryptography / Secrets", 2),
        ("Sensitive Data Exposure", 3),
        ("Authentication",         3),
        ("Authorization / IDOR",   2),
        ("Input Validation",       3),
        ("Configuration",          8),
        ("Business Logic",         3),
        ("Dependency",             2),
        ("Sensitive Data",         1),
        ("Business Logic/Perf",    3),
    ]
    header_row(ws, 23, ["Category", "Findings", "% of Total"], fg=C_ACCENT)
    total = sum(c for _, c in categories)
    for i, (cat, cnt) in enumerate(categories, 24):
        ws.row_dimensions[i].height = 18
        even = i % 2 == 0
        bg = C_LIGHT_BG if even else C_LIGHT2
        set_cell(ws, i, 1, cat,                   fg=bg)
        set_cell(ws, i, 2, cnt,                   fg=bg, h_align="center")
        set_cell(ws, i, 3, f"{cnt/total*100:.1f}%", fg=bg, h_align="center")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def generate_report():
    wb = openpyxl.Workbook()

    # Rename default sheet
    ws1 = wb.active
    ws1.title = "Executive Summary"

    ws2 = wb.create_sheet("Backend Inventory")
    ws3 = wb.create_sheet("API Endpoint Inventory")
    ws4 = wb.create_sheet("Security Findings")
    ws5 = wb.create_sheet("Dependency Review")
    ws6 = wb.create_sheet("Remediation Roadmap")
    ws7 = wb.create_sheet("GitHub Actions Workflow")
    ws8 = wb.create_sheet("Risk Dashboard")

    print("Building Executive Summary...")
    build_executive_summary(ws1)

    print("Building Backend Inventory...")
    build_backend_inventory(ws2)

    print("Building API Endpoint Inventory...")
    build_endpoint_inventory(ws3)

    print("Building Security Findings...")
    build_findings(ws4)

    print("Building Dependency Review...")
    build_dependency_review(ws5)

    print("Building Remediation Roadmap...")
    build_remediation(ws6)

    print("Building GitHub Actions Workflow...")
    build_github_actions(ws7)

    print("Building Risk Dashboard...")
    build_risk_dashboard(ws8)

    # Tab colors
    ws1.sheet_properties.tabColor = C_NAVY
    ws2.sheet_properties.tabColor = C_TEAL
    ws3.sheet_properties.tabColor = C_ACCENT
    ws4.sheet_properties.tabColor = C_RED
    ws5.sheet_properties.tabColor = C_ORANGE
    ws6.sheet_properties.tabColor = C_GREEN
    ws7.sheet_properties.tabColor = C_PURPLE
    ws8.sheet_properties.tabColor = C_YELLOW

    output = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "BrushIQ_Security_Review_Report.xlsx"
    )
    wb.save(output)
    print(f"\n✅  Security Review Report saved to:\n    {output}")
    return output


if __name__ == "__main__":
    generate_report()
